# -*- coding: utf-8 -*-
"""
============================================================================
 GENERADOR DE DATOS DE DEMOSTRACION (ficticios) — v2
============================================================================
Crea, en esta misma carpeta, una version FICTICIA de cada reporte crudo que
espera "generar_kpis.py":
    - Detalle de Pago Lote y CTS DEMO - Oficina Principal.xlsx
    - Detalle de Pago Lote y CTS DEMO - Tiendas.xlsx
    - Detalle CTS Individual DEMO.xlsx
    - Extornos DEMO.xlsx
    - Abono de Sueldos y CTS DEMO.xlsx
    - REQUERIMIENTOS BACKOFFICE DEMO.csv

Los nombres de columnas y la estructura (dos archivos "Oficina Principal" /
"Tiendas", secciones apiladas, encabezado en fila 3 del CSV, etc.) imitan el
formato real de un core bancario, pero TODOS los valores (usuarios, agencias,
empresas, montos, fechas) son inventados con un generador de numeros
aleatorios de semilla fija.

LA HISTORIA QUE CUENTAN ESTOS DATOS (a proposito, no es aleatoria del todo):
  - Mes 1 (piloto): solo 2 agencias "con convenio" arrancan Pago Lote, con
    muy pocos dias de actividad y la mayor parte del volumen ejecutado por
    personal de APOYO rotativo, no por el equipo fijo.
  - Mes 2 (atipico): avalancha de CTS que obliga a trabajar un sabado (mas
    dias de actividad que dias habiles del mes).
  - Mes 3: se incorporan mas agencias con convenio (ramp-up).
  - Mes 4: cobertura completa de las agencias con convenio; se vuelve el mes
    de referencia ("pico") para medir eficiencia.
  - Mes 5 (en curso): el reporte se corta a mitad de mes; ademas arrancan
    recien las Aperturas de CTS/Ahorros en lote.
  - Una agencia sin convenio ("AGENCIA 07") igual registra un par de pagos
    lote por su cuenta -> anomalia de universo, a proposito.
  - Dos agencias con convenio YA incorporadas ("AGENCIA 01", "AGENCIA 03")
    siguen procesando un pequeño residual por su cuenta -> fuga.

Sirve para probar "generar_kpis.py" y "generar_diapositivas.py" de punta a
punta, incluyendo las partes mas interesantes de la logica: deteccion de
mes piloto/atipico/en-curso, centralizacion real (eficacia vs cobertura vs
fuga) y la correccion de persona-dias cuando hay personal de apoyo rotativo.

Uso:  python generar_datos_demo.py
============================================================================
"""
import os
import csv
import random
import datetime as dt

import openpyxl
from openpyxl.styles import Font

CARPETA = os.path.dirname(os.path.abspath(__file__))
random.seed(42)

MESES = ["2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]
DIAS_POR_MES = {"2026-04": 30, "2026-05": 31, "2026-06": 30, "2026-07": 31, "2026-08": 31}
CORTE_MES_EN_CURSO = "2026-08"
DIA_CORTE_EN_CURSO = 18   # el reporte de Agosto se "descarga" a mitad de mes

USUARIOS_EQUIPO = ["ANALISTA1", "ASISTENTE1", "ASISTENTE2"]
USUARIOS_APOYO = ["APOYO01", "APOYO02"]        # rotativo: puede volver en cualquier mes
PERSONAL_AGENCIA = "PERSONAL DE AGENCIA"       # generico: no se usa para ningun calculo de equipo

AGENCIAS = [f"AGENCIA {i:02d}" for i in range(1, 39)]      # 38 agencias de la red
AGENCIAS_CONVENIO = ["AGENCIA 01", "AGENCIA 02", "AGENCIA 03",
                     "AGENCIA 04", "AGENCIA 05", "AGENCIA 06"]  # universo REAL de Pago Lote
AGENCIA_ANOMALIA = "AGENCIA 07"                 # sin convenio, pero opera Pago Lote igual
AGENCIAS_CON_FUGA = ["AGENCIA 01", "AGENCIA 03"]  # ya incorporadas, aun con residual
EMPRESAS = [f"EMPRESA DEMO {n} SAC" for n in
            ["ALFA", "BETA", "GAMMA", "DELTA", "OMEGA", "ZETA", "NORTE", "SUR"]]

# Que agencias con convenio ya estan incorporadas a Pago Lote, mes a mes
# (esto ES la cobertura de despliegue -- K6b la deberia reconstruir sola).
INCORPORADAS_POR_MES = {
    "2026-04": ["AGENCIA 01", "AGENCIA 02"],
    "2026-05": ["AGENCIA 01", "AGENCIA 02"],
    "2026-06": ["AGENCIA 01", "AGENCIA 02", "AGENCIA 03", "AGENCIA 04"],
    "2026-07": AGENCIAS_CONVENIO,
    "2026-08": AGENCIAS_CONVENIO,
}

# Dias de actividad de Oficina Principal por mes (a proposito, no todos los
# dias del mes): el mes 1 es un piloto de pocos dias; el mes 2 fuerza un
# sabado para quedar ATIPICO; el mes 5 se corta antes de DIA_CORTE_EN_CURSO.
def _dias_habiles_aprox(mes, hasta=None, margen=0):
    """
    Dias de semana (lun-vie) del mes. 'margen' descuenta los ultimos N dias de
    la lista a proposito: generar_kpis.py resta feriados nacionales reales al
    calcular dias habiles, y esta lista no los conoce -- descontar un margen
    evita que un mes que deberia ser NORMAL salga ATIPICO por 1-2 feriados
    que caigan justo en un dia que aca se listo como "habil".
    """
    anio, m = map(int, mes.split("-"))
    ultimo = hasta or DIAS_POR_MES[mes]
    dias = [dt.date(anio, m, d) for d in range(1, ultimo + 1) if dt.date(anio, m, d).weekday() < 5]
    return dias[:-margen] if margen else dias


DIAS_ACTIVIDAD_OP = {
    "2026-04": [dt.date(2026, 4, 6), dt.date(2026, 4, 7), dt.date(2026, 4, 13),
                dt.date(2026, 4, 14), dt.date(2026, 4, 20), dt.date(2026, 4, 27)],
    "2026-05": _dias_habiles_aprox("2026-05") + [dt.date(2026, 5, 9), dt.date(2026, 5, 16)],  # +2 sabados
    "2026-06": _dias_habiles_aprox("2026-06", margen=2),
    "2026-07": _dias_habiles_aprox("2026-07", margen=2),
    "2026-08": _dias_habiles_aprox("2026-08", hasta=DIA_CORTE_EN_CURSO - 1, margen=2),
}


def fecha_random(mes, dias_candidatos=None):
    anio, m = map(int, mes.split("-"))
    if dias_candidatos:
        dia = random.choice(dias_candidatos)
    else:
        dia = dt.date(anio, m, random.randint(1, DIAS_POR_MES[mes]))
    hora = random.randint(8, 17)
    minuto = random.randint(0, 59)
    return dt.datetime(dia.year, dia.month, dia.day, hora, minuto)


def escribir_hoja(ruta, secciones, con_cabecera_proceso=False):
    """
    secciones: lista de (encabezados, filas) que se escriben una tras otra.
    con_cabecera_proceso: si True, antepone 3 filas imitando el reporte real
    (fila en blanco, titulo, "Fecha de Proceso: dd/mm/aaaa") -- necesario para
    que fecha_proceso() en generar_kpis.py pueda detectar el corte del ultimo
    mes ("EN CURSO").
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if con_cabecera_proceso:
        hoy = dt.date(2026, 8, DIA_CORTE_EN_CURSO)
        ws.append([])
        ws.append(["", "", "", "", "Detalle de Pago Lote y CTS (DEMO)"])
        ws.append(["", "", f"Fecha de Proceso: {hoy.strftime('%d/%m/%Y')}"])
    for encabezados, filas in secciones:
        ws.append(encabezados)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for fila in filas:
            ws.append(fila)
        ws.append([])  # fila en blanco entre secciones
    wb.save(ruta)


# ---------------------------------------------------------------------------
# 1) Detalle de Pago Lote y CTS -- AHORA EN DOS ARCHIVOS (Oficina Principal /
#    Tiendas), igual que el reporte real desde que el core lo partio en dos.
# ---------------------------------------------------------------------------
def generar_detalle_pago_lote():
    encabezados = ["Numero de Movimiento", "Tipo de Operacion", "Monto",
                   "Fecha y Hora", "Usuario", "Tienda", "Empresa"]
    mov = 100000
    filas_op, filas_tie = [], []

    def usuario_del_mes(mes):
        # Mes 1: piloto con equipo chico y mucho apoyo. Mes 2: el apoyo se
        # retira salvo un ultimo apoyo puntual. Mes 5: reaparece un apoyo
        # (el apoyo es rotativo, puede volver en cualquier momento).
        if mes == "2026-04":
            return random.choices(USUARIOS_EQUIPO[:2] + USUARIOS_APOYO,
                                   weights=[10, 5, 45, 40])[0]
        if mes == "2026-05":
            return random.choices(USUARIOS_EQUIPO + USUARIOS_APOYO,
                                   weights=[10, 45, 40, 4, 1])[0]
        if mes == "2026-08":
            return random.choices(USUARIOS_EQUIPO + [USUARIOS_APOYO[0]],
                                   weights=[10, 45, 40, 5])[0]
        return random.choice(USUARIOS_EQUIPO)

    # --- PAGO LOTE: solo agencias CON CONVENIO ya incorporadas ese mes ---
    base_pago = {"2026-04": 20, "2026-05": 25, "2026-06": 55, "2026-07": 170, "2026-08": 75}
    for mes in MESES:
        incorporadas = INCORPORADAS_POR_MES[mes]
        n = base_pago[mes] + random.randint(-5, 5)
        for _ in range(n):
            mov += 1
            filas_op.append([mov, "PAGO LOTE", round(random.uniform(800, 15000), 2),
                              fecha_random(mes, DIAS_ACTIVIDAD_OP[mes]), usuario_del_mes(mes),
                              random.choice(incorporadas), random.choice(EMPRESAS)])
        # Fuga: agencias YA incorporadas que igual siguen procesando algo por
        # su cuenta (mas fuerte cuanto mas reciente la incorporacion).
        for ag in [a for a in AGENCIAS_CON_FUGA if a in incorporadas]:
            for _ in range(random.randint(0, 3)):
                mov += 1
                filas_tie.append([mov, "PAGO LOTE", round(random.uniform(800, 6000), 2),
                                   fecha_random(mes), PERSONAL_AGENCIA, ag, random.choice(EMPRESAS)])
        # Agencias con convenio TODAVIA NO incorporadas: lo siguen haciendo
        # ellas mismas (asi se arma la cobertura de despliegue, K6b).
        for ag in [a for a in AGENCIAS_CONVENIO if a not in incorporadas]:
            for _ in range(random.randint(3, 10)):
                mov += 1
                filas_tie.append([mov, "PAGO LOTE", round(random.uniform(800, 12000), 2),
                                   fecha_random(mes), PERSONAL_AGENCIA, ag, random.choice(EMPRESAS)])
        # Anomalia: una agencia SIN convenio que igual opera Pago Lote.
        if random.random() > 0.4:
            mov += 1
            filas_tie.append([mov, "PAGO LOTE", round(random.uniform(500, 3000), 2),
                               fecha_random(mes), PERSONAL_AGENCIA, AGENCIA_ANOMALIA,
                               random.choice(EMPRESAS)])

    # --- DEPOSITO LOTE CTS: universo amplio, avalancha en el mes 2 ---
    base_cts = {"2026-04": 15, "2026-05": 220, "2026-06": 60, "2026-07": 30, "2026-08": 20}
    for mes in MESES:
        n = base_cts[mes] + random.randint(-5, 5)
        for _ in range(n):
            mov += 1
            filas_op.append([mov, "DEPOSITO LOTE CTS", round(random.uniform(1500, 25000), 2),
                              fecha_random(mes, DIAS_ACTIVIDAD_OP[mes]), usuario_del_mes(mes),
                              random.choice(AGENCIAS[:25]), random.choice(EMPRESAS)])
        # Lo que las agencias siguen resolviendo de CTS por su cuenta.
        for _ in range(random.randint(20, 60)):
            mov += 1
            filas_tie.append([mov, "DEPOSITO LOTE CTS", round(random.uniform(1000, 15000), 2),
                               fecha_random(mes), PERSONAL_AGENCIA,
                               random.choice(AGENCIAS), random.choice(EMPRESAS)])

    # --- APERTURA CTS/AHORROS LOTE: arrancan recien en el ultimo mes ---
    for mes in MESES:
        if mes != "2026-08":
            # Antes de que el proceso exista en Of. Principal, las agencias
            # siguen abriendo cuentas cada una por su lado (cobertura baja).
            for _ in range(random.randint(15, 30)):
                mov += 1
                tipo = random.choice(["APERTURA CTS LOTE EFECTIVO", "APERTURA DE AHORROS LOTE EFECTIVO"])
                filas_tie.append([mov, tipo, round(random.uniform(0, 500), 2),
                                   fecha_random(mes), PERSONAL_AGENCIA,
                                   random.choice(AGENCIAS), random.choice(EMPRESAS)])
            continue
        for _ in range(random.randint(2, 5)):
            mov += 1
            filas_op.append([mov, "APERTURA CTS LOTE EFECTIVO", 0.0,
                              fecha_random(mes, DIAS_ACTIVIDAD_OP[mes]), usuario_del_mes(mes),
                              random.choice(AGENCIAS_CONVENIO), random.choice(EMPRESAS)])
        for _ in range(random.randint(10, 25)):
            mov += 1
            filas_op.append([mov, "APERTURA DE AHORROS LOTE EFECTIVO", 0.0,
                              fecha_random(mes, DIAS_ACTIVIDAD_OP[mes]), usuario_del_mes(mes),
                              random.choice(AGENCIAS_CONVENIO), random.choice(EMPRESAS)])
        # Aun asi, la mayoria de agencias las sigue haciendo por su cuenta
        # (cobertura de aperturas empieza en 0, como en el caso real).
        for _ in range(random.randint(10, 20)):
            mov += 1
            tipo = random.choice(["APERTURA CTS LOTE EFECTIVO", "APERTURA DE AHORROS LOTE EFECTIVO"])
            filas_tie.append([mov, tipo, round(random.uniform(0, 500), 2),
                               fecha_random(mes), PERSONAL_AGENCIA,
                               random.choice(AGENCIAS), random.choice(EMPRESAS)])

    # --- Registro de Pendientes por Devolver: trabajo estable del equipo,
    # se registra con Tienda = OFICINA PRINCIPAL (no es operacion de agencia).
    for mes in MESES:
        for _ in range(random.randint(8, 20)):
            mov += 1
            filas_op.append([mov, "REGISTRO DE PENDIENTES POR DEVOLVER CREDITOS CONVENIO",
                              round(random.uniform(200, 3000), 2),
                              fecha_random(mes, DIAS_ACTIVIDAD_OP[mes]), usuario_del_mes(mes),
                              "OFICINA PRINCIPAL", random.choice(EMPRESAS)])

    # --- Evento puntual sin monto: NO debe contarse en el volumen total
    # (asi se prueba que la familia "OTRO" se descarta correctamente).
    for _ in range(3):
        mov += 1
        filas_op.append([mov, "DEVOLUCION PAGO POR CONVENIO", 0.0,
                          fecha_random("2026-04", DIAS_ACTIVIDAD_OP["2026-04"]),
                          random.choice(USUARIOS_APOYO), "OFICINA PRINCIPAL", random.choice(EMPRESAS)])

    ruta_op = os.path.join(CARPETA, "Detalle de Pago Lote y CTS DEMO - Oficina Principal.xlsx")
    ruta_tie = os.path.join(CARPETA, "Detalle de Pago Lote y CTS DEMO - Tiendas.xlsx")
    escribir_hoja(ruta_op, [(encabezados, filas_op)], con_cabecera_proceso=True)
    escribir_hoja(ruta_tie, [(encabezados, filas_tie)])
    return ruta_op, ruta_tie


# ---------------------------------------------------------------------------
# 2) Detalle CTS Individual (por cliente)
# ---------------------------------------------------------------------------
def generar_detalle_cts():
    encabezados = ["N° de Movimiento", "Fecha", "Tienda", "Empresa",
                   "Tipo Operacion", "Monto"]
    filas = []
    mov = 500000
    tipos = ["APERTURA", "DEPOSITO", "CANCELACION"]
    for mes in MESES:
        n = random.randint(60, 100)
        for _ in range(n):
            mov += 1
            filas.append([mov, fecha_random(mes), random.choice(AGENCIAS[:20]),
                          random.choice(EMPRESAS),
                          random.choices(tipos, weights=[2, 7, 1])[0],
                          round(random.uniform(1000, 20000), 2)])
    ruta = os.path.join(CARPETA, "Detalle CTS Individual DEMO.xlsx")
    escribir_hoja(ruta, [(encabezados, filas)])
    return ruta


# ---------------------------------------------------------------------------
# 3) Extornos
# ---------------------------------------------------------------------------
def generar_extornos():
    encabezados = ["N° Mov Extorno", "Fecha Extorno", "Monto", "Motivo"]
    filas = []
    ext = 900000
    motivos = ["ERROR USUARIO", "ERROR SISTEMA", "ERROR CLIENTE"]
    for mes in MESES:
        n = random.randint(2, 6)
        for _ in range(n):
            ext += 1
            filas.append([ext, fecha_random(mes), round(random.uniform(500, 8000), 2),
                          random.choices(motivos, weights=[5, 3, 2])[0]])
    ruta = os.path.join(CARPETA, "Extornos DEMO.xlsx")
    escribir_hoja(ruta, [(encabezados, filas)])
    return ruta


# ---------------------------------------------------------------------------
# 4) Abono de Sueldos y CTS (toda la red)
# ---------------------------------------------------------------------------
def generar_abono():
    encabezados = ["Fecha Operacion", "Tipo Operacion", "Estado",
                   "Nombre Empleador", "Canal", "Tienda Operacion de Abono"]
    filas = []
    estados = ["EXITOSO", "OBSERVADO", "RECHAZADO"]
    for mes in MESES:
        # CTS resuelto en Oficina Principal (equipo centralizado) desde siempre.
        for _ in range(random.randint(80, 140)):
            filas.append([fecha_random(mes), "ABONO CTS",
                          random.choices(estados, weights=[9, 1, 0.3])[0],
                          random.choice(EMPRESAS), "LOTE", "OFICINA PRINCIPAL"])
        # El abono de SUELDO centralizado recien arranca el ultimo mes (igual
        # que las Aperturas): antes de eso, cada agencia hacia el suyo.
        if mes == CORTE_MES_EN_CURSO:
            for _ in range(random.randint(150, 220)):
                filas.append([fecha_random(mes), "ABONO SUELDO",
                              random.choices(estados, weights=[9, 1, 0.3])[0],
                              random.choice(EMPRESAS), "LOTE", "OFICINA PRINCIPAL"])
        # Resto de la red: sueldo y CTS resuelto de forma local.
        for _ in range(random.randint(300, 420)):
            filas.append([fecha_random(mes), random.choice(["ABONO SUELDO", "ABONO CTS"]),
                          random.choices(estados, weights=[9, 1, 0.3])[0],
                          random.choice(EMPRESAS), "LOTE", random.choice(AGENCIAS)])
    ruta = os.path.join(CARPETA, "Abono de Sueldos y CTS DEMO.xlsx")
    escribir_hoja(ruta, [(encabezados, filas)])
    return ruta


# ---------------------------------------------------------------------------
# 5) Buzon de Requerimientos (CSV ';' con encabezado en fila 3)
# ---------------------------------------------------------------------------
def generar_requerimientos():
    ruta = os.path.join(CARPETA, "REQUERIMIENTOS BACKOFFICE DEMO.csv")
    tipos_solicitud = ["APERTURA CUENTA", "ABONO CTS", "CONSULTA SALDO",
                       "RECLAMO", "ACTUALIZACION DATOS", "OTROS"]
    asignados = USUARIOS_EQUIPO
    with open(ruta, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["REQUERIMIENTOS BACKOFFICE (DATOS DE EJEMPLO)"])
        w.writerow([])
        w.writerow(["N°", "Fecha de Solicitud", "Tipo de Solicitud", "Asignado a",
                    "Fecha de Atencion", "Estado"])
        n = 1
        for mes in MESES:
            for _ in range(random.randint(30, 45)):
                f_sol = fecha_random(mes).date()
                tiene_fecha_atencion = random.random() > 0.35
                f_at = (f_sol + dt.timedelta(days=random.randint(0, 5))) if tiene_fecha_atencion else ""
                w.writerow([n, f_sol.strftime("%d/%m/%Y"),
                           random.choice(tipos_solicitud), random.choice(asignados),
                           f_at.strftime("%d/%m/%Y") if f_at else "",
                           "ATENDIDO" if tiene_fecha_atencion else "PENDIENTE"])
                n += 1
    return ruta


def main():
    print("Generando datos de demostracion (ficticios) en:", CARPETA)
    ruta_op, ruta_tie = generar_detalle_pago_lote()
    rutas = [ruta_op, ruta_tie, generar_detalle_cts(), generar_extornos(),
             generar_abono(), generar_requerimientos()]
    for r in rutas:
        print("  ->", os.path.basename(r))
    print("\nListo. Ahora puedes correr:  python generar_kpis.py")


if __name__ == "__main__":
    main()
