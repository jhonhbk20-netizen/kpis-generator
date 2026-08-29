# -*- coding: utf-8 -*-
"""
============================================================================
 GENERADOR DE KPIs — PAGOS MASIVOS Y CTS (v2)
============================================================================
Lee automaticamente los reportes crudos de Excel/CSV que dejes en ESTA
carpeta y genera un unico archivo Excel con todos los KPIs calculados,
listos para armar las diapositivas del analisis mensual.

COMO USARLO
-----------
1) Deja en esta carpeta los reportes crudos (pueden tener sufijos como
   "(1)", "(2)", copia, etc. — el script los detecta por su nombre):
       - Detalle de Pago Lote y CTS ... Oficina Principal.xlsx  (lo que
         ejecuta el equipo: Pago Lote, Deposito CTS, Aperturas, Pendientes)
       - Detalle de Pago Lote y CTS ... Tiendas.xlsx  (lo mismo, pero hecho
         por las agencias -- se usa para medir centralizacion y fuga)
       - Detalle CTS Individual ....xlsx          (CTS individual por cliente)
       - Extornos ....xlsx
       - Abono de Sueldos y CTS ....xlsx
       - REQUERIMIENTOS BACKOFFICE ....csv
   Si hay varios que coinciden, usa el MAS RECIENTE (fecha de modificacion).
   (No tienes estos reportes? Corre antes "generar_datos_demo.py" para crear
   una version ficticia de cada uno y probar el flujo completo.)
2) Ejecuta:  python generar_kpis.py
3) Se crea:  KPIs_Pagos_CTS_<MES>.xlsx  (con hoja de Explicacion).

No necesitas editar rutas. Si quieres fijar el mes destacado, cambia
MES_OBJETIVO abajo; si lo dejas en None se destaca el ultimo mes con datos.
============================================================================
"""

import os
import re
import json
import glob
import unicodedata
import datetime as dt
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.text import CharacterProperties

# ---------------------------------------------------------------------------
# CONFIGURACION (ajustable) --------------------------------------------------
# ---------------------------------------------------------------------------
CARPETA        = os.path.dirname(os.path.abspath(__file__))   # esta misma carpeta
MES_OBJETIVO   = None          # ej. "2026-07"; None = ultimo mes con datos
TOTAL_TIENDAS_RED = 38         # denominador de cobertura CTS: agencias de tu red

# --- Universo de Pago Lote: SOLO agencias con convenio -----------------------
# Pago Lote es cobranza por descuento en planilla: solo aplica a agencias que
# tienen convenio con un empleador. El resto de la red NO es un pendiente de
# despliegue -- no hay nada que centralizar ahi. Usar TOTAL_TIENDAS_RED aca
# da una cobertura falsa (pocas agencias / toda la red en vez de / convenio).
# Si se deja en None se deriva de la data (agencias con actividad de Pago
# Lote en Oficina Principal), asi un convenio nuevo entra solo.
TIENDAS_CONVENIO = None
# Agencias que registran una familia sin pertenecer a su universo, POR
# FAMILIA: en este demo, "AGENCIA 07" no tiene convenio pero igual aparece
# procesando Pago Lote por su cuenta -- se aparta del denominador y se
# reporta aparte como anomalia (revisar registro), sin tocar otras familias.
FUERA_UNIVERSO_POR_FAMILIA = {"PAGO": {"AGENCIA 07"}}

# --- Clasificacion de meses (define que meses son comparables) ---------------
# Un mes solo sirve de referencia de capacidad si fue un mes de trabajo
# normal: ni un piloto de pocos dias, ni un mes con sobretiempo, ni un
# reporte a medio cortar.
MIN_OPS_DIA_EFECTIVO = 5       # ops minimas para considerar que un dia tuvo volumen real
UMBRAL_MES_PARCIAL   = 0.60    # dias de actividad / dias habiles por debajo de esto = mes parcial

# Historico "congelado": guarda los KPIs ya calculados por mes para que, al
# soltar un mes nuevo, los anteriores no se vuelvan a recalcular -- solo se
# agregan los meses que todavia no estan aqui. Se pregunta en cada corrida.
CACHE_HISTORICO = os.path.join(CARPETA, "kpis_historico_demo_cache.json")

# Roles de operadores (usuario del core -> rol). Ajusta a tu propio equipo.
ROLES = {
    "ANALISTA1": "Analista",
    "ASISTENTE1": "Asistente 1",
    "ASISTENTE2": "Asistente 2",
}
ANALISTAS = {u for u, r in ROLES.items() if r.lower().startswith("analista")}
ASISTENTES = {u for u, r in ROLES.items() if r.lower().startswith("asistente")}

DIAS_SEMANA_ES = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

# Tienda que representa a nuestro equipo (backoffice) dentro del reporte de
# Abono Sueldos y CTS, que junta TODAS las agencias de la red.
TIENDA_PROPIA = "OFICINA PRINCIPAL"

# Supuestos de tiempo de procesamiento (ESTIMADO, no cronometrado)
MIN_POR_PAGO_LOTE = 1.3        # minutos por transaccion de Pago Lote
MIN_POR_CTS       = 4.0        # minutos por transaccion de CTS
HORAS_JORNADA     = 8.0        # horas por dia laboral (para pasar a dias-persona)


# ---------------------------------------------------------------------------
# UTILIDADES -----------------------------------------------------------------
# ---------------------------------------------------------------------------
def norm(texto):
    """minusculas + sin tildes + espacios colapsados (para comparar textos)."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def buscar_archivo(*claves, ext=("xlsx", "xls", "csv")):
    """Devuelve el archivo mas reciente cuyo nombre contenga TODAS las claves."""
    candidatos = []
    for e in ext:
        candidatos += glob.glob(os.path.join(CARPETA, f"*.{e}"))
    def coincide(nombre):
        base = os.path.basename(nombre)
        if base.startswith("~$"):   # archivo temporal de Office (esta abierto)
            return False
        low = base.lower()
        return all(k.lower() in low for k in claves)
    hits = [c for c in candidatos if coincide(c)]
    if not hits:
        return None
    return max(hits, key=os.path.getmtime)   # el mas reciente


def buscar_archivos(*claves, ext=("xlsx", "xls", "csv")):
    """Como buscar_archivo, pero devuelve TODOS los que coinciden (no solo el ultimo)."""
    candidatos = []
    for e in ext:
        candidatos += glob.glob(os.path.join(CARPETA, f"*.{e}"))
    def coincide(nombre):
        base = os.path.basename(nombre)
        if base.startswith("~$"):
            return False
        low = base.lower()
        return all(k.lower() in low for k in claves)
    return sorted([c for c in candidatos if coincide(c)], key=os.path.getmtime, reverse=True)


def mes_de(valor):
    """Convierte una fecha/valor a 'YYYY-MM' o None."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (dt.datetime, dt.date)):
        return f"{valor.year:04d}-{valor.month:02d}"
    ts = pd.to_datetime(valor, errors="coerce", dayfirst=True)
    return None if pd.isna(ts) else ts.strftime("%Y-%m")


def _domingo_pascua(anio):
    """Algoritmo de Meeus/Jones/Butcher para la fecha de Pascua (usada por feriados moviles)."""
    a = anio % 19
    b, c = divmod(anio, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes, dia = divmod(h + l - 7 * m + 114, 31)
    return dt.date(anio, mes, dia + 1)


def feriados_peru(anio):
    """Feriados nacionales fijos + moviles (Semana Santa) de un anio dado."""
    pascua = _domingo_pascua(anio)
    fijos = {dt.date(anio, 1, 1), dt.date(anio, 5, 1), dt.date(anio, 6, 29),
             dt.date(anio, 7, 28), dt.date(anio, 7, 29), dt.date(anio, 8, 30),
             dt.date(anio, 10, 8), dt.date(anio, 11, 1), dt.date(anio, 12, 8),
             dt.date(anio, 12, 25)}
    moviles = {pascua - dt.timedelta(days=2), pascua - dt.timedelta(days=3)}
    return fijos | moviles


def dias_habiles_mes(mes_str):
    """Dias habiles (lun-vie, sin feriados nacionales) de un mes 'YYYY-MM'."""
    anio, mes = (int(x) for x in mes_str.split("-"))
    fin = dt.date(anio + (mes == 12), (mes % 12) + 1, 1) - dt.timedelta(days=1)
    fer = feriados_peru(anio)
    dia, n = dt.date(anio, mes, 1), 0
    while dia <= fin:
        if dia.weekday() < 5 and dia not in fer:
            n += 1
        dia += dt.timedelta(days=1)
    return n


def fecha_de(valor):
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    ts = pd.to_datetime(valor, errors="coerce", dayfirst=True)
    return None if pd.isna(ts) else ts.date()


def a_numero(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def leer_filas(ruta):
    """Lee todas las filas de la primera hoja como listas (valores)."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas = list(wb.active.iter_rows(values_only=True))
    wb.close()
    return filas


def mapa_columnas(fila_encabezado):
    """De una fila de encabezado construye {nombre_normalizado: indice}."""
    m = {}
    for i, c in enumerate(fila_encabezado):
        if c is None:
            continue
        nombre = norm(c)
        if nombre and nombre not in m:
            m[nombre] = i
    return m


def col(m, *nombres):
    """Devuelve el indice de la primera columna que coincida (por 'contiene')."""
    for n in nombres:
        n = norm(n)
        if n in m:
            return m[n]
    for n in nombres:
        n = norm(n)
        for nombre, idx in m.items():
            if n in nombre:
                return idx
    return None


def buscar_columna_df(df, *nombres):
    """Version de col() para un DataFrame ya leido con pandas."""
    m = {norm(c): c for c in df.columns}
    for n in nombres:
        n = norm(n)
        if n in m:
            return m[n]
    for n in nombres:
        n = norm(n)
        for nombre, real in m.items():
            if n in nombre:
                return real
    return None


# ---------------------------------------------------------------------------
# LECTORES DE CADA FUENTE ----------------------------------------------------
# ---------------------------------------------------------------------------
def buscar_detalle():
    """
    Devuelve (ruta_oficina_principal, ruta_tiendas) del reporte de Pago Lote y
    CTS.

    OJO — el core exporta este reporte en DOS archivos complementarios (0
    movimientos y 0 usuarios en comun): uno de OFICINA PRINCIPAL (lo que
    ejecuta el equipo -- fuente de todos los KPIs de volumen/carga/eficiencia)
    y otro de TIENDAS (lo mismo hecho por las agencias -- solo se usa como
    denominador de centralizacion/fuga residual). NO se puede elegir por
    fecha de modificacion: si ambos se descargan juntos, el mas reciente
    podria ser cualquiera de los dos. Se discrimina por NOMBRE.
    """
    todos = buscar_archivos("detalle", "lote")
    op = tie = None
    legado = []
    for r in todos:
        n = norm(os.path.basename(r))
        if "oficina principal" in n:
            op = r if op is None else max(op, r, key=os.path.getmtime)
        elif "tienda" in n:
            tie = r if tie is None else max(tie, r, key=os.path.getmtime)
        else:
            legado.append(r)
    # Respaldo: formato antiguo (un solo archivo, sin sufijo) = Oficina Principal.
    if op is None and legado:
        op = max(legado, key=os.path.getmtime)
    return op, tie


def fecha_proceso(ruta):
    """Fecha de corte del reporte, leida de su cabecera ("Fecha de Proceso: dd/mm/aaaa")."""
    try:
        for f in leer_filas(ruta)[:8]:
            for c in f:
                t = norm(c)
                if "fecha de proceso" in t:
                    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", str(c))
                    if m:
                        d, mo, a = (int(x) for x in m.groups())
                        return dt.date(a, mo, d)
    except Exception:
        pass
    return None


def familia(tipo):
    """
    Familia de operacion a partir de 'Tipo de Operacion'.

    OJO CON EL ORDEN: "APERTURA CTS LOTE EFECTIVO" contiene la palabra "CTS",
    asi que una regla generica de CTS capturaria una apertura como si fuera un
    deposito. Las aperturas se evaluan primero.
    """
    t = norm(tipo)
    if "pago lote" in t:
        return "PAGO"
    if "apertura" in t:
        return "APERTURA_CTS" if "cts" in t else "APERTURA_AHORROS"
    if "cts" in t:
        return "CTS_DEPOSITO"
    if "pendientes" in t:
        return "PENDIENTES"
    return "OTRO"


# Familia -> segmento historico (compatibilidad con el K7 original de 2 categorias).
_SEGMENTO_DE_FAMILIA = {"PAGO": "PAGO", "CTS_DEPOSITO": "CTS"}


def leer_detalle_centrales(ruta, origen="CENTRALES"):
    """
    Archivo con sub-tablas apiladas (Pago Lote, Deposito Lote CTS, aperturas,
    pendientes por devolver...). Se detecta cada encabezado por la columna
    'Numero de movimiento' y se mapean columnas por NOMBRE.

    'origen' marca de que archivo viene cada fila ('CENTRALES' = Oficina
    Principal, 'TIENDAS' = agencias), para poder concatenar ambos sin
    ambiguedad al calcular centralizacion.
    """
    filas = leer_filas(ruta)
    encabezados = [i for i, f in enumerate(filas)
                   if any("numero de movimiento" in norm(c) for c in f)]
    encabezados.append(len(filas))  # centinela final
    recs = []
    for k in range(len(encabezados) - 1):
        h = encabezados[k]
        fin = encabezados[k + 1]
        m = mapa_columnas(filas[h])
        i_mov  = col(m, "numero de movimiento")
        i_tipo = col(m, "tipo de operacion", "tipo de  operacion", "tipo")
        i_mon  = col(m, "monto")
        i_fec  = col(m, "fecha y hora", "fecha")
        i_usu  = col(m, "usuario")
        i_tie  = col(m, "tienda")
        i_emp  = col(m, "empresa", "empleador")
        for f in filas[h + 1:fin]:
            if i_mov is None or i_mov >= len(f):
                continue
            mov = f[i_mov]
            if mov in (None, "") or "numero de movimiento" in norm(mov):
                continue
            tipo = str(f[i_tipo]).strip().upper() if i_tipo is not None and i_tipo < len(f) and f[i_tipo] else ""
            if not tipo:
                continue
            fam = familia(tipo)
            fval = f[i_fec] if (i_fec is not None and i_fec < len(f)) else None
            recs.append({
                "familia": fam,
                "segmento": _SEGMENTO_DE_FAMILIA.get(fam, "OTRO"),
                "origen": origen,
                "tipo": tipo,
                "mov": mov,
                "monto": a_numero(f[i_mon]) if (i_mon is not None and i_mon < len(f)) else 0.0,
                "mes": mes_de(fval),
                "dia": fecha_de(fval),
                "usuario": (str(f[i_usu]).strip() if (i_usu is not None and i_usu < len(f) and f[i_usu]) else None),
                "tienda": (str(f[i_tie]).strip() if (i_tie is not None and i_tie < len(f) and f[i_tie]) else None),
                "empleador": (str(f[i_emp]).strip() if (i_emp is not None and i_emp < len(f) and f[i_emp]) else None),
            })
    df = pd.DataFrame(recs)
    return df[df["mes"].notna()].copy()


def leer_detalle_cts(ruta):
    """CTS individual por cliente (APERTURA / DEPOSITO / CANCELACION)."""
    filas = leer_filas(ruta)
    hi = next((i for i, f in enumerate(filas)
               if any("de movimiento" in norm(c) for c in f)), None)
    if hi is None:
        return pd.DataFrame()
    m = mapa_columnas(filas[hi])
    i_mov = col(m, "de movimiento", "n° de movimiento")
    i_fec = col(m, "fecha")
    i_tie = col(m, "tienda")
    i_emp = col(m, "empresa", "empleador")
    i_tipo = col(m, "tipo operacion", "tipo operación", "tipo")
    i_mon = col(m, "monto")
    recs = []
    for f in filas[hi + 1:]:
        if i_mov is None or i_mov >= len(f) or f[i_mov] in (None, ""):
            continue
        fval = f[i_fec] if (i_fec is not None and i_fec < len(f)) else None
        recs.append({
            "mes": mes_de(fval),
            "tienda": str(f[i_tie]).strip() if (i_tie is not None and i_tie < len(f) and f[i_tie]) else None,
            "empleador": str(f[i_emp]).strip() if (i_emp is not None and i_emp < len(f) and f[i_emp]) else None,
            "tipo": str(f[i_tipo]).strip().upper() if (i_tipo is not None and i_tipo < len(f) and f[i_tipo]) else "",
            "monto": a_numero(f[i_mon]) if (i_mon is not None and i_mon < len(f)) else 0.0,
        })
    df = pd.DataFrame(recs)
    return df[df["mes"].notna()].copy() if not df.empty else df


def leer_extornos(rutas):
    """
    Un extorno = un 'N° Mov Extorno' unico (varias filas = un evento).

    Se reciben TODOS los archivos de extorno que haya en la carpeta (el core
    puede exportar mas de uno con nombre casi identico -- ej. toda la red vs.
    solo la oficina propia). Se leen todos y se filtra por tienda/usuario del
    equipo para no mezclar errores de otras agencias en esta tasa.
    """
    if isinstance(rutas, str):
        rutas = [rutas]
    recs, vistos = [], set()
    for ruta in rutas:
        filas = leer_filas(ruta)
        hi = next((i for i, f in enumerate(filas)
                   if any("mov extorno" in norm(c) for c in f)), None)
        if hi is None:
            continue
        m = mapa_columnas(filas[hi])
        i_ext = col(m, "mov extorno", "n° mov extorno")
        i_fec = col(m, "fecha extorno", "fecha")
        i_mon = col(m, "monto")
        i_mot = col(m, "motivo")
        for f in filas[hi + 1:]:
            if i_ext is None or i_ext >= len(f) or f[i_ext] in (None, ""):
                continue
            fval = f[i_fec] if (i_fec is not None and i_fec < len(f)) else None
            clave = (f[i_ext], fval, f[i_mon] if (i_mon is not None and i_mon < len(f)) else None)
            if clave in vistos:
                continue
            vistos.add(clave)
            recs.append({
                "extorno": f[i_ext],
                "mes": mes_de(fval),
                "monto": a_numero(f[i_mon]) if (i_mon is not None and i_mon < len(f)) else 0.0,
                "motivo": str(f[i_mot]).strip().upper() if (i_mot is not None and i_mot < len(f) and f[i_mot]) else "SIN MOTIVO",
            })
    df = pd.DataFrame(recs)
    return df[df["mes"].notna()].copy() if not df.empty else df


def leer_abono(ruta):
    """Abono de Sueldos/CTS. No tiene Monto; tiene columna Estado."""
    filas = leer_filas(ruta)
    hi = next((i for i, f in enumerate(filas)
               if any("fecha operaci" in norm(c) for c in f)), None)
    if hi is None:
        return pd.DataFrame()
    m = mapa_columnas(filas[hi])
    i_fec = col(m, "fecha operacion", "fecha operación", "fecha")
    i_tipo = col(m, "tipo operacion", "tipo operación", "tipo")
    i_est = col(m, "estado")
    i_emp = col(m, "nombre empleador", "empleador")
    i_can = col(m, "canal")
    i_tie = col(m, "tienda operacion de abono", "tienda")
    recs = []
    for f in filas[hi + 1:]:
        if i_fec is None or i_fec >= len(f) or f[i_fec] in (None, ""):
            continue
        recs.append({
            "mes": mes_de(f[i_fec]),
            "tipo": str(f[i_tipo]).strip().upper() if (i_tipo is not None and i_tipo < len(f) and f[i_tipo]) else "",
            "estado": str(f[i_est]).strip().upper() if (i_est is not None and i_est < len(f) and f[i_est]) else "",
            "empleador": str(f[i_emp]).strip() if (i_emp is not None and i_emp < len(f) and f[i_emp]) else None,
            "canal": str(f[i_can]).strip().upper() if (i_can is not None and i_can < len(f) and f[i_can]) else "",
            "tienda": str(f[i_tie]).strip().upper() if (i_tie is not None and i_tie < len(f) and f[i_tie]) else "",
        })
    df = pd.DataFrame(recs)
    return df[df["mes"].notna()].copy() if not df.empty else df


def leer_requerimientos(ruta):
    """Buzon de requerimientos (CSV separado por ';', encabezado en fila 3)."""
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            raw = pd.read_csv(ruta, sep=";", encoding=enc, header=None,
                              dtype=str, keep_default_na=False, engine="python")
            break
        except Exception:
            raw = None
    if raw is None or raw.empty:
        return pd.DataFrame()
    hi = next((i for i in range(min(8, len(raw)))
               if raw.iloc[i].astype(str).str.contains("SOLICITUD", case=False).any()
               and (raw.iloc[i].astype(str).str.strip() != "").sum() >= 5), 2)
    header = [re.sub(r"\s+", " ", str(c)).strip() for c in raw.iloc[hi]]
    data = raw.iloc[hi + 1:].copy()
    data.columns = header
    data = data.dropna(how="all")
    return data


# ---------------------------------------------------------------------------
# CLASIFICACION DE MESES -------------------------------------------------
# ---------------------------------------------------------------------------
def clasificar_meses(det, corte=None):
    """
    Etiqueta cada mes segun si es COMPARABLE con los demas. Sin esto, el pico
    de referencia de K7 se ancla a meses que no representan trabajo normal y
    todos los demas meses aparecen con una brecha de capacidad ficticia.

    Clases:
      EN CURSO  -> el reporte corta a mitad de mes (no esta cerrado todavia).
      PARCIAL   -> se trabajaron muy pocos dias frente a los habiles del mes
                   (ej. un piloto que arranca en pocas agencias).
      ATIPICO   -> mas dias de actividad que dias habiles = se trabajo sabado
                   o feriado (ej. una avalancha estacional).
      NORMAL    -> mes de trabajo regular. Solo estos son comparables.
    """
    info = {}
    if det.empty:
        return info
    d = det.dropna(subset=["dia"])
    for mes in sorted(det["mes"].dropna().unique()):
        dm = d[d["mes"] == mes]
        por_dia = dm.groupby("dia").size()
        dias_act = int(len(por_dia))
        dias_efec = int((por_dia >= MIN_OPS_DIA_EFECTIVO).sum())
        dias_hab = dias_habiles_mes(mes) or 0

        en_curso = bool(corte and mes == corte.strftime("%Y-%m")
                        and corte < ultimo_dia_habil(mes))
        if en_curso:
            clase = "EN CURSO"
            nota = f"Reporte cortado al {corte.strftime('%d/%m')} - mes incompleto"
        elif dias_hab and dias_act / dias_hab < UMBRAL_MES_PARCIAL:
            clase = "PARCIAL"
            nota = (f"Solo {dias_act} dias de actividad de {dias_hab} habiles"
                    + (f" ({dias_efec} con volumen real)" if dias_efec < dias_act else ""))
        elif dias_act > dias_hab > 0:
            clase = "ATIPICO"
            nota = f"{dias_act} dias de actividad vs {dias_hab} habiles (sobretiempo)"
        else:
            clase, nota = "NORMAL", ""

        info[mes] = {
            "clase": clase, "nota": nota,
            "dias_actividad": dias_act, "dias_efectivos": dias_efec,
            "dias_habiles": dias_hab,
            "comparable": clase == "NORMAL",
        }
    return info


def ultimo_dia_habil(mes_str):
    """Ultimo dia habil (lun-vie, sin feriados) del mes 'YYYY-MM'."""
    a, m = (int(x) for x in mes_str.split("-"))
    fin = dt.date(a + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1)
    fer = feriados_peru(a)
    while fin.weekday() >= 5 or fin in fer:
        fin -= dt.timedelta(days=1)
    return fin


# ---------------------------------------------------------------------------
# CALCULO DE KPIs ------------------------------------------------------------
# ---------------------------------------------------------------------------
def _metricas_mes(det, cts, mes):
    """Todo lo que las distintas tablas de un mes necesitan, calculado una sola vez."""
    d = det[det["mes"] == mes]
    # familia "OTRO" (eventos puntuales sin monto, ej. una devolucion aislada)
    # queda FUERA del volumen: no es una familia de trabajo continua.
    d_ops = d[d["familia"] != "OTRO"]
    pago = d_ops[d_ops["segmento"] == "PAGO"]
    ctsl = d_ops[d_ops["segmento"] == "CTS"]

    ops_usu = d_ops.groupby("usuario").size().to_dict()
    asist = {u: n for u, n in ops_usu.items() if u in ASISTENTES}
    k5 = (round(max(asist.values()) / max(min(asist.values()), 1), 2)
          if len(asist) >= 2 else np.nan)
    analista_ops = sum(n for u, n in ops_usu.items() if u in ANALISTAS)
    total_ops = len(d_ops)

    cts_mes = cts[cts["mes"] == mes] if not cts.empty else pd.DataFrame()
    fuente_cts = cts_mes if not cts_mes.empty else ctsl
    tiendas_cts = fuente_cts["tienda"].dropna().nunique()
    empleadores = fuente_cts["empleador"].dropna().nunique()

    # OJO — el denominador de productividad NO se filtra por lista de
    # usuarios. Si se filtrara a ROLES mientras el numerador cuenta a todos,
    # el trabajo del personal de apoyo se le atribuiria a la plantilla fija.
    # El apoyo es rotativo -- puede volver en cualquier momento -- asi que se
    # cuentan las persona-dias de TODOS los que trabajaron; ROLES queda solo
    # para K5 (balance entre asistentes de plantilla fija).
    dd = d_ops.dropna(subset=["dia", "usuario"])
    dd_apoyo = dd[~dd["usuario"].isin(ROLES)]
    persona_dias = dd.groupby(["usuario", "dia"]).ngroups
    dias_actividad = dd["dia"].nunique()
    dias_calendario = dias_habiles_mes(mes)
    operadores_activos = dd["usuario"].nunique()

    return {
        "mes": mes, "d": d_ops, "pago": pago, "ctsl": ctsl, "ops_usu": ops_usu,
        "k1": len(pago), "k2": pago["monto"].sum(),
        "k3": len(ctsl), "k4": ctsl["monto"].sum(),
        "ap_cts": d_ops[d_ops["familia"] == "APERTURA_CTS"],
        "ap_aho": d_ops[d_ops["familia"] == "APERTURA_AHORROS"],
        "pendient": d_ops[d_ops["familia"] == "PENDIENTES"],
        "total_ops": total_ops, "total_monto": d_ops["monto"].sum(),
        "k5": k5, "k5a": analista_ops / total_ops if total_ops else 0.0,
        "tiendas_pago": pago["tienda"].dropna().nunique(),
        "tiendas_cts": tiendas_cts, "empleadores": empleadores,
        "persona_dias": persona_dias,
        "persona_dias_fijo": dd[dd["usuario"].isin(ROLES)].groupby(["usuario", "dia"]).ngroups,
        "persona_dias_apoyo": dd_apoyo.groupby(["usuario", "dia"]).ngroups,
        "ops_apoyo": len(dd_apoyo),
        "dias_actividad": dias_actividad, "dias_calendario": dias_calendario,
        "operadores_activos": operadores_activos,
        "ops_dia_persona": total_ops / persona_dias if persona_dias else 0,
        "ops_dia_activo": total_ops / dias_actividad if dias_actividad else 0,
        "ops_dia_calend": (total_ops / (operadores_activos * dias_calendario)
                           if operadores_activos and dias_calendario else 0),
    }


def _fila_resumen(m):
    """
    Solo KPIs de verdad -- valores que se reportan y se comparan mes a mes.
    Operadores/Persona-dias/Dias con actividad/Ops-dia NO van aca: son
    variables de calculo, no un KPI en si, y ya estan completas -y
    explicadas- en la seccion K7.
    """
    return {
        "Mes": m["mes"], "K1 Vol Pago": m["k1"], "K2 Monto Pago": round(m["k2"], 2),
        "K3 Vol CTS": m["k3"], "K4 Monto CTS": round(m["k4"], 2),
        "Vol Apertura CTS": len(m["ap_cts"]),
        "Vol Apertura Ahorros/Sueldo": len(m["ap_aho"]),
        "Vol Pendientes x Devolver": len(m["pendient"]),
        "Monto Pendientes (S/)": round(m["pendient"]["monto"].sum(), 2),
        "Total Ops": m["total_ops"], "Total Monto": round(m["total_monto"], 2),
        "K5 Ratio Asist.": "N/A" if pd.isna(m["k5"]) else m["k5"],
        "K5a % Analista": round(m["k5a"], 4),
        "Empleadores CTS": m["empleadores"],
    }


def _filas_operadores(m):
    """Desglosa Total Ops en sus familias -- para ver de donde sale cada persona."""
    filas = []
    for u in sorted([x for x in m["ops_usu"] if x]):
        du = m["d"][m["d"]["usuario"] == u]
        dpago = du[du["segmento"] == "PAGO"]
        dcts = du[du["segmento"] == "CTS"]
        dap = du[du["familia"].isin(("APERTURA_CTS", "APERTURA_AHORROS"))]
        dpend = du[du["familia"] == "PENDIENTES"]
        dias_u = du["dia"].dropna().nunique()
        filas.append({
            "Mes": m["mes"], "Usuario": u, "Rol": ROLES.get(u, "Apoyo/Otro"),
            "Ops Pago": len(dpago), "Monto Pago": round(dpago["monto"].sum(), 2),
            "Ops CTS": len(dcts), "Monto CTS": round(dcts["monto"].sum(), 2),
            "Ops Apertura": len(dap), "Ops Pendientes x Devolver": len(dpend),
            "Total Ops": len(du), "Total Monto": round(du["monto"].sum(), 2),
            "Dias Activos": dias_u,
            "Ops/Dia": round(len(du) / dias_u, 1) if dias_u else 0,
        })
    return filas


def _fila_carga_mensual(m):
    """
    "Dia Pico" = el dia con mas operaciones del mes. "Dias Pico en el Mes" =
    cuantos dias superaron promedio + 1 desviacion estandar de ese mes.
    "Movimientos Pago Lote" es una APROXIMACION a corridas/cheques (agrupa
    por Nro de Movimiento), NO un conteo exacto: algunas agencias registran
    los pagos partidos en varios movimientos.
    """
    d_dia = m["d"].dropna(subset=["dia"])
    if d_dia.empty:
        return {"Mes": m["mes"], "Dia Pico": "", "Dia Semana Pico": "",
                "Ops Dia Pico": 0, "Movimientos Pago Lote (dia pico)": 0,
                "Dias Pico en el Mes": 0, "Dia Semana Mas Cargado (prom.)": ""}
    por_dia = d_dia.groupby("dia").size()
    sigma = por_dia.std(ddof=0)
    umbral = por_dia.mean() + (sigma if sigma else 0)
    dia_pico = por_dia.idxmax()
    dow = d_dia["dia"].apply(lambda f: DIAS_SEMANA_ES[f.weekday()])
    promedio_dow = d_dia.groupby(dow).size() / d_dia.groupby(dow)["dia"].nunique()
    return {
        "Mes": m["mes"], "Dia Pico": dia_pico.isoformat(),
        "Dia Semana Pico": DIAS_SEMANA_ES[dia_pico.weekday()],
        "Ops Dia Pico": int(por_dia.max()),
        "Movimientos Pago Lote (dia pico)":
            int(m["pago"][m["pago"]["dia"] == dia_pico]["mov"].dropna().nunique()),
        "Dias Pico en el Mes": int((por_dia >= umbral).sum()),
        "Dia Semana Mas Cargado (prom.)": promedio_dow.idxmax() if not promedio_dow.empty else "",
    }


def _fila_eficiencia(m, meses_info):
    clase = (meses_info or {}).get(m["mes"], {})
    return {
        "Mes": m["mes"], "Clase": clase.get("clase", "NORMAL"),
        "Nota Clase": clase.get("nota", ""),
        "Total Ops": m["total_ops"], "Operadores Activos": m["operadores_activos"],
        "Persona-dias Reales": m["persona_dias"],
        "Persona-dias Plantilla": m["persona_dias_fijo"],
        "Persona-dias Apoyo": m["persona_dias_apoyo"],
        "Ops de Apoyo": m["ops_apoyo"],
        "Dias con Actividad": m["dias_actividad"],
        "Dias Efectivos": clase.get("dias_efectivos", m["dias_actividad"]),
        "Dias Calendario": m["dias_calendario"],
        "Ops/Dia (persona-dia)": round(m["ops_dia_persona"], 1),
        "Ops/Dia (dia-activo)": round(m["ops_dia_activo"], 1),
        "Ops/Dia (calendario)": round(m["ops_dia_calend"], 1),
        "Brecha": round(m["ops_dia_persona"] - m["ops_dia_calend"], 1),
    }


def _fila_tiempo(m):
    """Estimado de carga horaria. Tabla auxiliar (no es un KPI numerado)."""
    horas_pago = m["k1"] * MIN_POR_PAGO_LOTE / 60
    horas_cts = m["k3"] * MIN_POR_CTS / 60
    horas_tot = horas_pago + horas_cts
    return {
        "Mes": m["mes"], "K1 Pago Lote": m["k1"], "K3 CTS": m["k3"],
        "Horas Pago Lote": round(horas_pago, 1), "Horas CTS": round(horas_cts, 1),
        "Horas Totales Estim.": round(horas_tot, 1),
        "Equiv. Dias (8h)": round(horas_tot / HORAS_JORNADA, 1),
    }


def _fila_extornos(m, ext):
    """El denominador es el TOTAL de operaciones del mes, no solo Pago Lote."""
    e = ext[ext["mes"] == m["mes"]] if not ext.empty else pd.DataFrame()
    if e.empty:
        return None

    def ev(motivo):
        return e[e["motivo"].str.contains(motivo)]["extorno"].nunique()

    tot_ext, imp, tot = e["extorno"].nunique(), ev("USUARIO"), m["total_ops"]
    return {
        "Mes": m["mes"], "K1 Pago Lote": m["k1"], "K3 CTS": m["k3"],
        "Vol Apertura CTS": len(m["ap_cts"]), "Vol Apertura Ahorros/Sueldo": len(m["ap_aho"]),
        "Vol Pendientes x Devolver": len(m["pendient"]),
        "Total Ops (denominador)": tot,
        "Extornos Totales": tot_ext, "Imputables (Error Usuario)": imp,
        "Error Sistema": ev("SISTEMA"), "Error Cliente": ev("CLIENTE"),
        "Tasa Imputable": round(imp / tot, 5) if tot else 0,
        "Tasa Total": round(tot_ext / tot, 5) if tot else 0,
        "Monto Extornado (S/)": round(e["monto"].sum(), 2),
    }


def _filas_k9(det, abo, meses):
    """
    K9: aperturas + abonos de SUELDO, TODO medido sobre el trabajo DEL EQUIPO
    (filtrado a TIENDA_PROPIA). El abono de CTS no va aca porque es el mismo
    dato que K3 del resumen -- reportarlo aca seria redundante.
    """
    filas = []
    for mes in meses:
        d = det[det["mes"] == mes]
        ap_cts = int((d["familia"] == "APERTURA_CTS").sum())
        ap_aho = int((d["familia"] == "APERTURA_AHORROS").sum())
        fila = {"Mes": mes, "Aperturas CTS": ap_cts,
                "Aperturas Ahorros/Sueldo": ap_aho,
                "Total Aperturas": ap_cts + ap_aho}
        if not abo.empty and mes in set(abo["mes"].dropna()):
            a_red = abo[(abo["mes"] == mes) & (abo["tipo"].str.contains("SUELDO"))]
            a = a_red[a_red["tienda"] == TIENDA_PROPIA]
            exito = int(a["estado"].str.contains("EXITOSO").sum()) if len(a) else 0
            fila.update({
                "Abonos Sueldo": len(a),
                "% Abonos Exitosos": round(exito / len(a), 4) if len(a) else 0,
                "Empleadores": int(a["empleador"].dropna().nunique()),
                "Abonos Sueldo Red (contexto)": len(a_red),
                "% Abonos Centralizados": round(len(a) / len(a_red), 4) if len(a_red) else 0,
                "Nota": "",
            })
        else:
            fila.update({
                "Abonos Sueldo": None, "% Abonos Exitosos": None, "Empleadores": None,
                "Abonos Sueldo Red (contexto)": None, "% Abonos Centralizados": None,
                "Nota": "Sin datos: el reporte de Abono no cubre este mes",
            })
        filas.append(fila)
    return filas


def calcular(det, cts, ext, abo, meses_incluir=None, meses_info=None):
    """
    Arma todas las tablas de KPIs, un mes por vez.

    meses_incluir: si se pasa (lista/set de 'YYYY-MM'), solo se calculan esos
        meses -- se usa en modo historico para procesar SOLO los meses nuevos.
    meses_info: salida de clasificar_meses(); marca que meses son comparables.
    """
    meses = sorted(det["mes"].dropna().unique().tolist()) if not det.empty else []
    if meses_incluir is not None:
        meses = [m for m in meses if m in set(meses_incluir)]

    resumen, operadores, extornos_k8 = [], [], []
    tiempo, eficiencia_k7, carga_mensual = [], [], []

    for mes in meses:
        m = _metricas_mes(det, cts, mes)
        resumen.append(_fila_resumen(m))
        operadores.extend(_filas_operadores(m))
        carga_mensual.append(_fila_carga_mensual(m))
        eficiencia_k7.append(_fila_eficiencia(m, meses_info))
        tiempo.append(_fila_tiempo(m))
        fila_ext = _fila_extornos(m, ext)
        if fila_ext:
            extornos_k8.append(fila_ext)

    return dict(resumen=resumen, operadores=operadores, extornos=extornos_k8,
                tiempo=tiempo, eficiencia=eficiencia_k7, carga_mensual=carga_mensual,
                k9=_filas_k9(det, abo, meses), meses=meses)


def sintetizar_eficiencia(eficiencia_k7):
    """
    Brecha de capacidad: usa el ritmo mensual mas alto observado ENTRE MESES
    NORMAL (Ops/Dia persona-dia) como referencia de "capacidad al 100%". Un
    mes PARCIAL, ATIPICO o EN CURSO no puede ser la vara: fijaria una
    capacidad que el equipo no sostiene y todos los demas meses apareceria
    con una brecha que no existe.
    """
    if not eficiencia_k7:
        return eficiencia_k7
    ritmos_exactos = {f["Mes"]: (f["Total Ops"] / f["Persona-dias Reales"]
                                  if f["Persona-dias Reales"] else 0)
                       for f in eficiencia_k7}
    comparables = {m: r for m, r in ritmos_exactos.items()
                   if next((f.get("Clase", "NORMAL") == "NORMAL"
                            for f in eficiencia_k7 if f["Mes"] == m), True)}
    base = comparables or ritmos_exactos
    ritmo_max = max(base.values()) if base else 0
    mes_pico = max(base, key=base.get) if base else ""
    for f in eficiencia_k7:
        ritmo = ritmos_exactos[f["Mes"]]
        pdias = f["Persona-dias Reales"]
        comparable = f.get("Clase", "NORMAL") == "NORMAL"
        f["Mes Pico Ref."] = mes_pico
        f["Ritmo vs Pico (%)"] = round(ritmo / ritmo_max, 4) if ritmo_max else 0
        f["Brecha Capacidad (ops)"] = round((ritmo_max - ritmo) * pdias) if comparable else ""
        if not comparable:
            detalle = f.get("Nota Clase", "")
            f["Nota"] = f"Mes {f.get('Clase','')} - no comparable" + (f": {detalle}" if detalle else "")
        elif f["Persona-dias Apoyo"]:
            f["Nota"] = (f"Incluye {f['Ops de Apoyo']} ops de personal de apoyo "
                         f"({f['Persona-dias Apoyo']} persona-dias)")
        else:
            f["Nota"] = ""
        f.pop("Nota Clase", None)
    return eficiencia_k7


# ---------------------------------------------------------------------------
# CENTRALIZACION REAL (Oficina Principal vs Tiendas) -------------------------
# ---------------------------------------------------------------------------
# (clave, etiqueta, tiene_cobertura)
# tiene_cobertura=False para operaciones que NO son de agencia: "Pendientes
# por Devolver" se registra con tienda = OFICINA PRINCIPAL, asi que contar
# "agencias incorporadas" ahi no significa nada. Su eficacia si aplica.
FAMILIAS_CENTRALIZABLES = [
    ("PAGO",             "Pago Lote",                  True),
    ("CTS_DEPOSITO",     "Deposito CTS",               True),
    ("APERTURA_CTS",     "Apertura CTS",               True),
    ("APERTURA_AHORROS", "Apertura Ahorros/Sueldo",    True),
    ("PENDIENTES",       "Reg. Pendientes x Devolver", False),
]


def calcular_centralizacion(det_op, det_tie):
    """
    Centralizacion real, cruzando lo que hace el equipo (Oficina Principal)
    contra lo que las agencias siguen haciendo por su cuenta (archivo
    Tiendas). Devuelve TRES tablas -- un solo numero de "centralizacion"
    mezclaria tres preguntas distintas:

      eficacia  (K6a) - de lo que esta DENTRO del alcance, cuanto lo hace el
                        equipo.
      cobertura (K6b) - cuantas agencias del universo YA se incorporaron.
      fuga      (K6c) - operaciones que la agencia sigue haciendo POR SU
                        CUENTA pese a estar ya incorporada.

    DOS REGLAS, sin las cuales el numero sale falso:
      1) Corte por entrada en produccion: no se cuenta a las agencias antes
         de que exista el proceso para esa familia.
      2) Denominador = universo real de la familia, y solo agencias ya
         incorporadas. Para Pago Lote son las agencias CON CONVENIO, no toda
         la red: las que no tienen convenio no tienen nada que centralizar.
    """
    vacio = dict(eficacia=[], cobertura=[], fuga=[], anomalias=[], universo={})
    if det_op is None or det_op.empty:
        return vacio
    if det_tie is None or det_tie.empty:
        det_tie = det_op.iloc[0:0]

    eficacia, cobertura, fuga, anomalias, universos = [], [], [], [], {}

    for clave, etiqueta, tiene_cobertura in FAMILIAS_CENTRALIZABLES:
        op = det_op[det_op["familia"] == clave].dropna(subset=["dia"])
        tie = det_tie[det_tie["familia"] == clave].dropna(subset=["dia"])
        if op.empty:
            continue

        excluidas = FUERA_UNIVERSO_POR_FAMILIA.get(clave, set())
        fuera = tie[tie["tienda"].str.upper().isin(excluidas)] if (not tie.empty and excluidas) else tie.iloc[0:0]
        tie = tie[~tie["tienda"].str.upper().isin(excluidas)] if (not tie.empty and excluidas) else tie
        for tienda, g in (fuera.groupby("tienda") if not fuera.empty else []):
            anomalias.append({
                "Familia": etiqueta, "Tienda": tienda, "Operaciones": len(g),
                "Monto (S/)": round(g["monto"].sum(), 2),
                "Dias": ", ".join(sorted({d.isoformat() for d in g["dia"]})),
                "Nota": "Opera sin pertenecer al universo - revisar registro",
            })

        inicio = op["dia"].min()
        universo = (set(TIENDAS_CONVENIO) if (clave == "PAGO" and TIENDAS_CONVENIO)
                    else set(op["tienda"].dropna()) | set(tie["tienda"].dropna()))
        universos[etiqueta] = sorted(universo)

        alta = op.dropna(subset=["tienda"]).groupby("tienda")["dia"].min().to_dict()
        meses = sorted(set(op["mes"]) | set(tie[tie["dia"] >= inicio]["mes"] if not tie.empty else []))
        incorporadas = set()
        for mes in meses:
            om = op[op["mes"] == mes]
            tm = tie[tie["mes"] == mes] if not tie.empty else tie
            nuevas = sorted({t for t, d in alta.items() if d.strftime("%Y-%m") == mes})
            incorporadas |= set(nuevas)

            res = tm[tm["tienda"].isin(incorporadas)] if not tm.empty else tm
            a, b = len(om), len(res)
            eficacia.append({
                "Mes": mes, "Familia": etiqueta,
                "Ops Centrales": a, "Ops en Tienda (ya incorporadas)": b,
                "Total en Alcance": a + b,
                "% Centralizacion": round(a / (a + b), 4) if (a + b) else 0,
            })
            if not tiene_cobertura:
                continue
            cobertura.append({
                "Mes": mes, "Familia": etiqueta,
                "Tiendas Incorporadas": len(incorporadas),
                "Universo": len(universo),
                "% Cobertura": round(len(incorporadas) / len(universo), 4) if universo else 0,
                "Nuevas del Mes": ", ".join(nuevas),
                "Pendientes": ", ".join(sorted(universo - incorporadas)) or "-",
            })

        if not tie.empty:
            for tienda, g in tie.groupby("tienda"):
                if tienda not in alta:
                    continue
                g = g[g["dia"] >= alta[tienda]]
                if g.empty:
                    continue
                por_mes = g.groupby("mes").size()
                ult = meses[-1] if meses else None
                prev = meses[-2] if len(meses) > 1 else None
                u = int(por_mes.get(ult, 0)) if ult else 0
                p = int(por_mes.get(prev, 0)) if prev else 0
                if u == 0:
                    tendencia = "cerrada"
                elif prev is None:
                    tendencia = "nueva"
                elif u > p:
                    tendencia = "SUBE"
                elif u < p:
                    tendencia = "baja"
                else:
                    tendencia = "estable"
                fuga.append({
                    "Familia": etiqueta, "Tienda": tienda,
                    "Incorporada": alta[tienda].isoformat(),
                    "Ops en Tienda (post-alta)": int(por_mes.sum()),
                    "Monto (S/)": round(g["monto"].sum(), 2),
                    "Ultimo Mes": u, "Mes Anterior": p, "Tendencia": tendencia,
                })

    activas = [f for f in fuga if f["Ultimo Mes"] > 0]
    cerradas = [f for f in fuga if f["Ultimo Mes"] == 0]
    activas.sort(key=lambda r: (-r["Ultimo Mes"], -r["Ops en Tienda (post-alta)"]))
    resumen_cerradas = []
    for _, etiqueta, _cob in FAMILIAS_CENTRALIZABLES:
        c = [f for f in cerradas if f["Familia"] == etiqueta]
        if c:
            resumen_cerradas.append({
                "Familia": etiqueta, "Tiendas con Fuga Cerrada": len(c),
                "Ops Historicas": sum(f["Ops en Tienda (post-alta)"] for f in c),
                "Monto Historico (S/)": round(sum(f["Monto (S/)"] for f in c), 2),
            })
    return dict(eficacia=eficacia, cobertura=cobertura, fuga=activas,
                fuga_cerradas=resumen_cerradas, anomalias=anomalias,
                universo=universos)


def fusionar_absorcion_en_resumen(K):
    """
    Agrega al Resumen la pregunta que de verdad importa ahi: "de las
    operaciones de Pago Lote / CTS, cuantas las hace el EQUIPO y cuantas las
    siguen haciendo las agencias que ya se unieron" -- NO cobertura
    geografica (K6b), sino absorcion/eficacia de lo que ya esta dentro del
    alcance (K6a). Se fusiona por mes y familia para no duplicar el calculo.
    """
    eficacia = (K.get("centralizacion_real") or {}).get("eficacia", [])
    abs_pago = {f["Mes"]: f["% Centralizacion"] for f in eficacia if f["Familia"] == "Pago Lote"}
    abs_cts = {f["Mes"]: f["% Centralizacion"] for f in eficacia if f["Familia"] == "Deposito CTS"}
    for fila in K["resumen"]:
        fila["% Absorcion Pago Lote (Equipo)"] = abs_pago.get(fila["Mes"])
        fila["% Absorcion CTS (Equipo)"] = abs_cts.get(fila["Mes"])


def calcular_buzon(req):
    """Demanda del buzon de requerimientos: por tipo de solicitud y por persona asignada."""
    buzon_tipo, buzon_asig, buzon_alerta = [], [], None
    if req.empty:
        return buzon_tipo, buzon_asig, buzon_alerta
    def col_req(*nombres):
        cols = {norm(c): c for c in req.columns}
        for n in nombres:
            if norm(n) in cols:
                return cols[norm(n)]
        for n in nombres:
            for k, v in cols.items():
                if norm(n) in k:
                    return v
        return None
    c_sol = col_req("solicitud")
    c_asig = col_req("asignado a", "asignado")
    c_fat = col_req("fecha de atencion")
    total = len(req)
    def limpio(serie):
        return (serie.fillna("").astype(str).str.strip()
                .str.replace(r"\s+", " ", regex=True)
                .str.upper().replace("", np.nan).dropna())
    if c_sol:
        vc = limpio(req[c_sol]).value_counts()
        buzon_tipo = [{"Tipo de Solicitud": k, "Cantidad": int(v),
                       "% del Total": round(v / total, 4)} for k, v in vc.items()]
    if c_asig:
        vc = limpio(req[c_asig]).value_counts()
        buzon_asig = [{"Asignado a": k, "Cantidad": int(v),
                       "% del Total": round(v / total, 4)} for k, v in vc.items()]
    if c_fat is not None:
        sin_fecha = req[c_fat].fillna("").str.strip().eq("").sum()
        buzon_alerta = (f"BRECHA: {sin_fecha} de {total} solicitudes SIN fecha de "
                        f"atencion -> el tiempo de respuesta del buzon hoy es invisible.")
    return buzon_tipo, buzon_asig, buzon_alerta


# ---------------------------------------------------------------------------
# CARGA LABORAL · DIAS PICO (detalle dia por dia, todo el historico crudo) --
# ---------------------------------------------------------------------------
def calcular_carga_diaria(det):
    """
    Detalle dia por dia de TODO el archivo crudo disponible (se recalcula
    siempre sobre el 'det' que se acaba de leer, no depende del modo
    historico). "Pico" = el dia supera promedio + 1 desviacion estandar del
    total de operaciones de ESE mes.
    """
    d = det.dropna(subset=["dia"])
    if d.empty:
        return {"dias": [], "por_dia_semana": []}
    d = d[d["familia"] != "OTRO"]

    diario = d.groupby("dia").agg(
        Ops_Pago=("familia", lambda s: int((s == "PAGO").sum())),
        Ops_CTS=("familia", lambda s: int((s == "CTS_DEPOSITO").sum())),
        Ops_Apertura=("familia", lambda s: int(s.isin(("APERTURA_CTS", "APERTURA_AHORROS")).sum())),
        Ops_Pendientes=("familia", lambda s: int((s == "PENDIENTES").sum())),
        Monto=("monto", "sum"),
        Mes=("mes", "first"),
    ).reset_index()
    diario["Total_Ops"] = diario["Ops_Pago"] + diario["Ops_CTS"] + diario["Ops_Apertura"] + diario["Ops_Pendientes"]
    diario["Dia Semana"] = diario["dia"].apply(lambda f: DIAS_SEMANA_ES[f.weekday()])

    umbral_por_mes = {}
    for mes, g in diario.groupby("Mes"):
        media, sigma = g["Total_Ops"].mean(), g["Total_Ops"].std(ddof=0)
        umbral_por_mes[mes] = media + sigma if sigma else media

    pago = det[det["segmento"] == "PAGO"]
    movs_por_dia = pago.dropna(subset=["dia"]).groupby("dia")["mov"].nunique()

    filas = []
    for _, r in diario.sort_values("dia").iterrows():
        es_pico = r["Total_Ops"] >= umbral_por_mes.get(r["Mes"], 0) and r["Total_Ops"] > 0
        filas.append({
            "Mes": r["Mes"], "Dia": r["dia"], "Dia Semana": r["Dia Semana"],
            "Ops Pago": int(r["Ops_Pago"]), "Ops CTS": int(r["Ops_CTS"]),
            "Ops Apertura": int(r["Ops_Apertura"]), "Ops Pendientes x Devolver": int(r["Ops_Pendientes"]),
            "Total Ops": int(r["Total_Ops"]),
            "Movimientos Pago Lote": int(movs_por_dia.get(r["dia"], 0)),
            "Monto Total": round(r["Monto"], 2),
            "Pico": "SI" if es_pico else "",
        })

    resumen_semana = []
    for dow in DIAS_SEMANA_ES:
        g = diario[diario["Dia Semana"] == dow]
        if g.empty:
            continue
        dias_pico_dow = sum(1 for _, r in g.iterrows()
                             if r["Total_Ops"] >= umbral_por_mes.get(r["Mes"], 0) and r["Total_Ops"] > 0)
        resumen_semana.append({
            "Dia Semana": dow, "Dias en el Periodo": len(g),
            "Total Ops (periodo)": int(g["Total_Ops"].sum()),
            "Promedio Ops/Dia": round(g["Total_Ops"].mean(), 1),
            "Dias Pico": dias_pico_dow,
        })
    resumen_semana.sort(key=lambda r: -r["Promedio Ops/Dia"])

    return {"dias": filas, "por_dia_semana": resumen_semana}


# ---------------------------------------------------------------------------
# HISTORICO "CONGELADO" · modo incremental por mes ---------------------------
# ---------------------------------------------------------------------------
_CAMPOS_HISTORICO = ("resumen", "operadores", "extornos", "tiempo",
                     "eficiencia", "carga_mensual", "k9")


def preguntar_si_no(pregunta):
    """input() con default 'no' si la entrada no es interactiva (ej. cron)."""
    try:
        resp = input(f"  {pregunta} (s/n): ").strip().lower()
    except EOFError:
        resp = ""
    return resp.startswith("s")


def cargar_historico():
    if not os.path.exists(CACHE_HISTORICO):
        return None
    try:
        with open(CACHE_HISTORICO, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError):
        print(f"  AVISO: no se pudo leer {os.path.basename(CACHE_HISTORICO)}, se ignora.")
        return None


def guardar_historico(K, meses_info=None):
    """Un mes EN CURSO NO se congela: quedaria fijado con datos incompletos."""
    excluir = {m for m, i in (meses_info or {}).items() if i["clase"] == "EN CURSO"}
    def limpiar(lista):
        return [f for f in lista if f.get("Mes") not in excluir]
    data = {campo: limpiar(K[campo]) for campo in _CAMPOS_HISTORICO}
    data["meses"] = [m for m in K["meses"] if m not in excluir]
    with open(CACHE_HISTORICO, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    if excluir:
        print(f"  (no se congelo {', '.join(sorted(excluir))}: mes aun en curso)")


def fusionar_historico(historico, K_nuevo):
    def combinar(lista_vieja, lista_nueva):
        meses_nuevos = {f["Mes"] for f in lista_nueva}
        return sorted([f for f in lista_vieja if f["Mes"] not in meses_nuevos] + lista_nueva,
                      key=lambda f: f["Mes"])
    K = {campo: combinar(historico.get(campo, []), K_nuevo[campo]) for campo in _CAMPOS_HISTORICO}
    K["eficiencia"] = sintetizar_eficiencia(K["eficiencia"])
    K["meses"] = sorted({f["Mes"] for f in K["resumen"]})
    return K


def reconstituir_desde_historico(historico):
    K = {campo: historico.get(campo, []) for campo in _CAMPOS_HISTORICO}
    K["eficiencia"] = sintetizar_eficiencia(K["eficiencia"])
    K["meses"] = sorted({f["Mes"] for f in K["resumen"]})
    return K


# ---------------------------------------------------------------------------
# ESCRITURA DEL EXCEL --------------------------------------------------------
# ---------------------------------------------------------------------------
AZUL   = "1F4E78"
CLARO  = "D9E1F2"
AMBAR  = "FFF2CC"
VERDE  = "E2EFDA"
GRIS   = "F2F2F2"
BORDE  = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _titulo(ws, texto, ncols):
    ws.append([texto])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(ncols, 1))
    c = ws.cell(r, 1)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22


def _tabla(ws, filas, mes_obj=None, pct_cols=(), money_cols=(), formulas=None, totales=()):
    """
    Escribe una tabla. 'formulas' escribe FORMULAS de Excel en vez de valores
    (asi el libro queda auditable: se ve de donde sale cada numero). 'totales'
    agrega una fila TOTAL con =SUMA() al pie. La hoja es compartida entre
    varias tablas, asi que el ancho de columna solo se ENSANCHA, nunca se
    encoge.
    """
    if not filas:
        ws.append(["(sin datos)"]); return
    cols = list(filas[0].keys())
    ws.append(cols)
    hr = ws.max_row
    for j in range(len(cols)):
        c = ws.cell(hr, j + 1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE

    formulas = formulas or {}
    primera = hr + 1
    for fila in filas:
        ws.append([fila.get(c) for c in cols])
        r = ws.max_row
        def ref(nombre, _r=r):
            return f"{get_column_letter(cols.index(nombre) + 1)}{_r}"
        es_obj = mes_obj and str(fila.get("Mes")) == mes_obj
        for j, cname in enumerate(cols):
            c = ws.cell(r, j + 1)
            c.border = BORDE
            if cname in formulas:
                try:
                    c.value = formulas[cname](ref)
                except (ValueError, KeyError):
                    pass
            if es_obj:
                c.fill = PatternFill("solid", fgColor=AMBAR)
            elif r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GRIS)
            if cname in pct_cols:
                c.number_format = "0.0%"
            elif cname in money_cols:
                c.number_format = "#,##0.00"
    ultima = ws.max_row

    if totales:
        ws.append([])
        ws.append(["TOTAL"] + [None] * (len(cols) - 1))
        r = ws.max_row
        for j, cname in enumerate(cols):
            c = ws.cell(r, j + 1)
            c.border = BORDE
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=CLARO)
            if cname in totales:
                letra = get_column_letter(j + 1)
                c.value = f"=SUM({letra}{primera}:{letra}{ultima})"
                if cname in money_cols:
                    c.number_format = "#,##0.00"

    for j, cname in enumerate(cols):
        largo = max([len(str(cname))] + [len(str(f.get(cname, ""))) for f in filas])
        ancho = min(max(largo + 2, 10), 40)
        dim = ws.column_dimensions[get_column_letter(j + 1)]
        dim.width = max(dim.width or 0, ancho)


def _nota(ws, texto):
    ws.append([texto])
    ws.cell(ws.max_row, 1).font = Font(italic=True, size=9, color="595959")
    ws.append([])


def _corte(ws):
    """Separador entre secciones apiladas en la hoja unica de calculos."""
    ws.append([]); ws.append([])


# Paleta para series de graficos (distinta de los colores de tabla: necesita
# buen contraste en barras/lineas). Pago=azul, CTS=verde, alerta=rojo,
# acento/meta=naranja.
COLOR_PAGO   = "1F4E78"
COLOR_CTS    = "2E8B57"
COLOR_ALERTA = "C00000"
COLOR_ACENTO = "ED7D31"
COLOR_GRIS2  = "808080"


def _estilo_serie(serie, color, linea=False, marcador=True):
    serie.graphicalProperties.solidFill = color
    serie.graphicalProperties.line.solidFill = color
    if linea:
        serie.graphicalProperties.line.width = 25000  # EMU ~ 2pt
        serie.smooth = False
        if marcador:
            serie.marker = Marker(symbol="circle", size=6)
            serie.marker.graphicalProperties.solidFill = color
            serie.marker.graphicalProperties.line.solidFill = color


def _eje(axis, texto):
    """Titulo de eje en negrita/legible (el default de openpyxl es muy chico)."""
    axis.title = texto
    try:
        axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=1100, b=True)
    except (AttributeError, IndexError):
        pass


def _hoja_graficos(wb, K):
    """Hoja 'Graficos': una tabla fuente compacta + graficos nativos de Excel."""
    ws = wb.create_sheet("Graficos")
    ws.sheet_view.showGridLines = False
    _titulo(ws, "GRAFICOS — PAGOS MASIVOS Y CTS", 13)

    resumen = K["resumen"]
    cob_por_mes = {(c["Mes"], c["Familia"]): c["% Cobertura"]
                   for c in (K.get("centralizacion_real") or {}).get("cobertura", [])}
    efic_por_mes = {e["Mes"]: e for e in K["eficiencia"]}

    fa_header = ["Mes", "K1 Vol Pago", "K3 Vol CTS", "K2 Monto Pago", "K4 Monto CTS",
                 "K5a %Analista", "K6b %Cobert Pago", "K6b %Cobert CTS", "K5 Ratio Asist.",
                 "Ops/dia persona", "Ops/dia calendario",
                 "K6a %Absorcion Pago Lote", "K6a %Absorcion CTS"]
    fa_start = ws.max_row + 1
    ws.append(fa_header)
    for j in range(len(fa_header)):
        ws.cell(fa_start, j + 1).font = Font(bold=True)
    for fila in resumen:
        k5 = fila["K5 Ratio Asist."]
        efic = efic_por_mes.get(fila["Mes"], {})
        ws.append([
            fila["Mes"], fila["K1 Vol Pago"], fila["K3 Vol CTS"],
            fila["K2 Monto Pago"], fila["K4 Monto CTS"], fila["K5a % Analista"],
            cob_por_mes.get((fila["Mes"], "Pago Lote")), cob_por_mes.get((fila["Mes"], "Deposito CTS")),
            None if k5 == "N/A" else k5,
            efic.get("Ops/Dia (persona-dia)"), efic.get("Ops/Dia (calendario)"),
            fila["% Absorcion Pago Lote (Equipo)"], fila["% Absorcion CTS (Equipo)"],
        ])
    fa_end = ws.max_row
    for col_pct in (6, 7, 8, 12, 13):
        for r in range(fa_start + 1, fa_end + 1):
            ws.cell(r, col_pct).number_format = "0.0%"
    ws.append([])

    fb_start = ws.max_row + 1
    ws.append(["Mes", "Tasa Imputable", "Tasa Total"])
    for j in range(3):
        ws.cell(fb_start, j + 1).font = Font(bold=True)
    for fila in K["extornos"]:
        ws.append([fila["Mes"], fila["Tasa Imputable"], fila["Tasa Total"]])
    fb_end = ws.max_row
    for r in range(fb_start + 1, fb_end + 1):
        ws.cell(r, 2).number_format = "0.0%"
        ws.cell(r, 3).number_format = "0.0%"
    ws.append([])

    fc_start = ws.max_row + 1
    ws.append(["Mes", "Horas Pago Lote", "Horas CTS"])
    for j in range(3):
        ws.cell(fc_start, j + 1).font = Font(bold=True)
    for fila in K["tiempo"]:
        ws.append([fila["Mes"], fila["Horas Pago Lote"], fila["Horas CTS"]])
    fc_end = ws.max_row

    for col in range(1, len(fa_header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    def refs(min_col, max_col, start, end):
        return (Reference(ws, min_col=min_col, max_col=max_col, min_row=start, max_row=end),
                Reference(ws, min_col=1, min_row=start + 1, max_row=end))

    CHART_H, CHART_W = 8.0, 15.0
    ALTO_FILA_CM, ANCHO_COL_CM = 0.53, 1.59
    ALTO = int((CHART_H + 2.5) / ALTO_FILA_CM) + 1
    ANCHO = int((CHART_W + 2.0) / ANCHO_COL_CM) + 1

    ANCLA_COL = len(fa_header) + 2
    fila_ancla = 3
    for col in range(ANCLA_COL, ANCLA_COL + 2 * ANCHO + 2):
        ws.column_dimensions[get_column_letter(col)].width = 8.43

    def ubicar(idx):
        fila = fila_ancla + (idx // 2) * ALTO
        col = ANCLA_COL + (idx % 2) * ANCHO
        return f"{get_column_letter(col)}{fila}"

    charts = []

    c1 = BarChart(); c1.type = "col"; c1.grouping = "clustered"; c1.style = 10
    c1.title = "Volumen de Transacciones por Proceso"
    _eje(c1.y_axis, "N° de transacciones"); _eje(c1.x_axis, "Mes")
    data, cats = refs(2, 3, fa_start, fa_end)
    c1.add_data(data, titles_from_data=True); c1.set_categories(cats)
    _estilo_serie(c1.series[0], COLOR_PAGO); _estilo_serie(c1.series[1], COLOR_CTS)
    charts.append(c1)

    c2 = BarChart(); c2.type = "col"; c2.grouping = "clustered"; c2.style = 10
    c2.title = "Monto Operado por Proceso (S/)"
    _eje(c2.y_axis, "Monto (S/)"); _eje(c2.x_axis, "Mes")
    data, cats = refs(4, 5, fa_start, fa_end)
    c2.add_data(data, titles_from_data=True); c2.set_categories(cats)
    _estilo_serie(c2.series[0], COLOR_PAGO); _estilo_serie(c2.series[1], COLOR_CTS)
    charts.append(c2)

    c3 = LineChart(); c3.style = 12
    c3.title = "K5a: % Transacciones Ejecutadas por el Analista"
    _eje(c3.y_axis, "% del total de ops"); _eje(c3.x_axis, "Mes")
    data, cats = refs(6, 6, fa_start, fa_end)
    c3.add_data(data, titles_from_data=True); c3.set_categories(cats)
    c3.y_axis.numFmt = "0%"
    _estilo_serie(c3.series[0], COLOR_ALERTA, linea=True)
    charts.append(c3)

    c4 = LineChart(); c4.style = 12
    c4.title = "K6b: Cobertura de Despliegue (universo real por familia)"
    _eje(c4.y_axis, "% de agencias del universo"); _eje(c4.x_axis, "Mes")
    data, cats = refs(7, 8, fa_start, fa_end)
    c4.add_data(data, titles_from_data=True); c4.set_categories(cats)
    c4.y_axis.numFmt = "0%"
    _estilo_serie(c4.series[0], COLOR_PAGO, linea=True)
    _estilo_serie(c4.series[1], COLOR_CTS, linea=True)
    charts.append(c4)

    c5 = LineChart(); c5.style = 12
    c5.title = "Eficiencia: Ops/Operador/Dia (persona-dia vs calendario)"
    _eje(c5.y_axis, "Ops por operador y dia"); _eje(c5.x_axis, "Mes")
    data, cats = refs(10, 11, fa_start, fa_end)
    c5.add_data(data, titles_from_data=True); c5.set_categories(cats)
    _estilo_serie(c5.series[0], COLOR_PAGO, linea=True)
    _estilo_serie(c5.series[1], COLOR_GRIS2, linea=True)
    charts.append(c5)

    c6 = BarChart(); c6.type = "col"; c6.style = 10
    c6.title = "K5: Ratio de Carga entre Asistentes"
    _eje(c6.y_axis, "Ratio MAX/MIN (meta <= 2.0x)"); _eje(c6.x_axis, "Mes")
    data, cats = refs(9, 9, fa_start, fa_end)
    c6.add_data(data, titles_from_data=True); c6.set_categories(cats)
    _estilo_serie(c6.series[0], COLOR_ACENTO)
    charts.append(c6)

    c7 = LineChart(); c7.style = 12
    c7.title = "K6a: % Absorcion — Pago Lote y CTS (dentro del alcance)"
    _eje(c7.y_axis, "% ejecutado por el equipo"); _eje(c7.x_axis, "Mes")
    data, cats = refs(12, 13, fa_start, fa_end)
    c7.add_data(data, titles_from_data=True); c7.set_categories(cats)
    c7.y_axis.numFmt = "0%"
    _estilo_serie(c7.series[0], COLOR_PAGO, linea=True)
    _estilo_serie(c7.series[1], COLOR_CTS, linea=True)
    charts.append(c7)

    if fb_end > fb_start:
        c8 = LineChart(); c8.style = 12
        c8.title = "K8: Tasa de Extornos (Imputable vs Total)"
        _eje(c8.y_axis, "% del total de operaciones"); _eje(c8.x_axis, "Mes")
        data, cats = refs(2, 3, fb_start, fb_end)
        c8.add_data(data, titles_from_data=True); c8.set_categories(cats)
        c8.y_axis.numFmt = "0.0%"
        _estilo_serie(c8.series[0], COLOR_ALERTA, linea=True)
        _estilo_serie(c8.series[1], COLOR_GRIS2, linea=True)
        charts.append(c8)

    if fc_end > fc_start:
        c9 = BarChart(); c9.type = "col"; c9.grouping = "stacked"; c9.style = 10
        c9.title = "Horas Estimadas por Proceso (auxiliar)"
        _eje(c9.y_axis, "Horas estimadas"); _eje(c9.x_axis, "Mes")
        data, cats = refs(2, 3, fc_start, fc_end)
        c9.add_data(data, titles_from_data=True); c9.set_categories(cats)
        _estilo_serie(c9.series[0], COLOR_PAGO); _estilo_serie(c9.series[1], COLOR_CTS)
        charts.append(c9)

    for idx, ch in enumerate(charts):
        ch.height, ch.width = CHART_H, CHART_W
        ch.legend.position = "b"
        ws.add_chart(ch, ubicar(idx))

    _nota(ws, "Graficos nativos de Excel a partir de las tablas de la izquierda. "
              "Se pueden copiar/pegar como imagen o editar el estilo directo en Excel.")
    return ws


def _sec_resumen(ws, K, mes_obj):
    _titulo(ws, "TABLERO DE KPIs — PAGOS MASIVOS Y CTS", 13)
    _nota(ws, f"Mes destacado: {mes_obj or '—'}  ·  Generado: {dt.datetime.now():%Y-%m-%d %H:%M}")
    _tabla(ws, K["resumen"], mes_obj,
           pct_cols=("K5a % Analista", "% Absorcion Pago Lote (Equipo)", "% Absorcion CTS (Equipo)"),
           money_cols=("K2 Monto Pago", "K4 Monto CTS", "Total Monto"),
           formulas={
               "Total Ops": lambda r: (
                   f"={r('K1 Vol Pago')}+{r('K3 Vol CTS')}+{r('Vol Apertura CTS')}"
                   f"+{r('Vol Apertura Ahorros/Sueldo')}+{r('Vol Pendientes x Devolver')}"),
           },
           totales=("K1 Vol Pago", "K3 Vol CTS", "Vol Apertura CTS",
                    "Vol Apertura Ahorros/Sueldo", "Vol Pendientes x Devolver",
                    "Total Ops", "K2 Monto Pago", "K4 Monto CTS", "Total Monto"))
    ws.append([])
    _nota(ws, "Las celdas de Total Ops y los porcentajes son FORMULAS de Excel, no valores pegados: "
              "se ve de donde sale cada numero y el libro recalcula si se corrige un dato a mano.")
    _nota(ws, "K1/K2: Pago Lote. K3/K4: Deposito Lote CTS.")
    _nota(ws, "K5 = MAX/MIN ops de asistentes. K5a = ops del analista / total.")
    _nota(ws, "% Absorcion = de lo que YA esta dentro del alcance (agencias incorporadas), cuanto lo "
              "ejecuta el equipo vs. cuanto lo sigue haciendo la agencia por su cuenta. NO es cobertura "
              "geografica (esa es K6b) -- es el mismo dato que K6a, mostrado aca junto al volumen del mes.")
    _nota(ws, "Operadores, Persona-dias, Dias con actividad/calendario y Ops/dia NO estan en esta tabla: "
              "son variables de calculo, no KPIs en si -- estan completas y explicadas en la seccion K7.")


def _sec_operadores(ws, K, mes_obj):
    _corte(ws)
    _titulo(ws, "DETALLE POR OPERADOR", 12)
    _tabla(ws, K["operadores"], mes_obj,
           money_cols=("Monto Pago", "Monto CTS", "Total Monto"))


def _sec_eficiencia(ws, K, mes_obj):
    """
    K7 se muestra en DOS tablas: una resumen (lo que se lee de un vistazo) y
    una de detalle metodologico (para quien quiera revisar de donde sale
    cada numero) -- juntarlas en una sola tabla de ~19 columnas parecidas
    entre si es dificil de leer.
    """
    _corte(ws)
    _titulo(ws, "K7 · EFICIENCIA OPERATIVA", 7)
    _nota(ws, "Persona-dias = cuantas persona-dias trabajo el equipo ese mes (incluye al personal de "
              "apoyo, ver detalle abajo). Ops/dia = Total Ops / Persona-dias: el ritmo real del equipo.")
    _nota(ws, "Clase del mes: NORMAL = comparable. PARCIAL = pocos dias trabajados frente a los habiles "
              "(ej. un piloto). ATIPICO = se trabajo sabado/feriado (ej. una avalancha estacional). "
              "EN CURSO = reporte cortado a mitad de mes. Solo los meses NORMAL sirven de vara.")
    _nota(ws, "Ritmo vs Pico = ritmo del mes / ritmo del 'Mes Pico Ref.' (el NORMAL mas alto). "
              "Brecha Capacidad = cuantas operaciones faltarian para igualar ese ritmo; queda vacia en "
              "meses no comparables, donde no significa nada.")
    cols_resumen = ("Mes", "Clase", "Total Ops", "Operadores Activos", "Dias con Actividad",
                    "Ops/Dia (persona-dia)", "Mes Pico Ref.",
                    "Ritmo vs Pico (%)", "Brecha Capacidad (ops)", "Nota")
    _tabla(ws, [{c: f[c] for c in cols_resumen} for f in K["eficiencia"]], mes_obj,
           pct_cols=("Ritmo vs Pico (%)",))
    ws.append([])

    _titulo(ws, "K7 · DETALLE METODOLOGICO (de donde sale cada numero de arriba)", 8)
    _nota(ws, "Plantilla/Apoyo = como se reparten las Persona-dias Reales entre el equipo fijo y el "
              "personal de apoyo rotativo (puede volver en cualquier momento).")
    _nota(ws, "Dia-activo = Total Ops / dias con actividad (ignora cuanta gente trabajo). Util en meses "
              "parciales, donde el ritmo por calendario hace parecer ocioso a un equipo que en realidad "
              "solo trabajo unos pocos dias por diseno (ej. un piloto).")
    _nota(ws, "Calendario = Total Ops / (operadores activos x dias habiles del mes). Responde 'que ritmo "
              "habria si todos trabajaran todos los dias habiles' -- un techo teorico, no el ritmo real.")
    cols_detalle = ("Mes", "Persona-dias Plantilla", "Persona-dias Apoyo", "Ops de Apoyo",
                    "Dias con Actividad", "Dias Calendario",
                    "Ops/Dia (dia-activo)", "Ops/Dia (calendario)", "Brecha")
    _tabla(ws, [{c: f[c] for c in cols_detalle} for f in K["eficiencia"]], mes_obj)
    ws.append([])

    _titulo(ws, "CARGA LABORAL · RESUMEN MENSUAL DE DIAS PICO", 7)
    _nota(ws, "Dia Pico = dia con mas operaciones del mes. Dias Pico en el Mes = dias que superaron el "
              "promedio + 1 desviacion estandar del mes.")
    _nota(ws, "Movimientos Pago Lote (dia pico) es una APROXIMACION a corridas/cheques (agrupa por Nro "
              "de Movimiento), NO un conteo exacto: algunas agencias registran pagos partidos.")
    _tabla(ws, K["carga_mensual"], mes_obj)
    ws.append([])

    CD = K.get("carga_diaria") or {}
    if CD.get("dias"):
        _titulo(ws, "CARGA LABORAL · RESUMEN POR DIA DE SEMANA", 5)
        _tabla(ws, CD["por_dia_semana"])
        ws.append([])

        _titulo(ws, "CARGA LABORAL · DETALLE DIA POR DIA (TODOS LOS DIAS DE ACTIVIDAD)", 9)
        _nota(ws, "Pico = el dia supera el promedio + 1 desviacion estandar del total de operaciones de ESE mes.")
        _tabla(ws, CD["dias"], money_cols=("Monto Total",))


def _sec_extornos(ws, K, mes_obj):
    _corte(ws)
    _titulo(ws, "K8 · TASA DE EXTORNOS", 10)
    _nota(ws, "Un extorno = un N° Mov Extorno unico. Imputable = motivo 'ERROR USUARIO'. "
              "Tasa = extornos / TOTAL de operaciones del mes (Pago Lote + CTS + aperturas + pendientes).")
    _tabla(ws, K["extornos"], mes_obj,
           pct_cols=("Tasa Imputable", "Tasa Total"),
           money_cols=("Monto Extornado (S/)",),
           formulas={
               "Tasa Imputable": lambda r: (
                   f"=IF({r('Total Ops (denominador)')}=0,\"\","
                   f"{r('Imputables (Error Usuario)')}/{r('Total Ops (denominador)')})"),
               "Tasa Total": lambda r: (
                   f"=IF({r('Total Ops (denominador)')}=0,\"\","
                   f"{r('Extornos Totales')}/{r('Total Ops (denominador)')})"),
           },
           totales=("Extornos Totales", "Imputables (Error Usuario)",
                    "Error Sistema", "Error Cliente", "Monto Extornado (S/)"))


def _sec_k9(ws, K, mes_obj):
    _corte(ws)
    _titulo(ws, "K9 · APERTURAS Y ABONOS", 10)
    _nota(ws, "Aperturas: conteo por familia del Detalle (tipos APERTURA...).")
    _nota(ws, f"Abonos: filtrados a lo que hace el EQUIPO (tienda = '{TIENDA_PROPIA}'), igual que las "
              "aperturas -- el total de la red queda como columna de contexto para comparar.")
    _tabla(ws, K["k9"], mes_obj, pct_cols=("% Abonos Exitosos", "% Abonos Centralizados"))
    ws.append([])
    _titulo(ws, "AUXILIAR · TIEMPO DE PROCESAMIENTO (ESTIMADO)", 7)
    _nota(ws, f"ADVERTENCIA: estimado, NO cronometrado. Supuesto: {MIN_POR_PAGO_LOTE} min/Pago Lote y {MIN_POR_CTS} min/CTS.")
    _tabla(ws, K["tiempo"], mes_obj)


def _sec_centralizacion(ws, K, mes_obj):
    CE = K.get("centralizacion_real") or {}
    if not CE.get("eficacia"):
        return
    _corte(ws)
    _titulo(ws, "K6 · CENTRALIZACION — EQUIPO vs AGENCIAS", 6)
    _nota(ws, "Fuente: cruce del archivo de OFICINA PRINCIPAL (lo que hace el equipo) contra el de "
              "TIENDAS (lo que las agencias siguen haciendo).")
    _nota(ws, "DOS REGLAS: (1) no se cuenta a las agencias antes de que el proceso entre en produccion "
              "para esa familia; (2) el denominador es el universo real de la familia y solo las "
              "agencias ya incorporadas -- no toda la red.")

    _titulo(ws, "K6a · EFICACIA DE CENTRALIZACION (dentro del alcance)", 6)
    _nota(ws, "De lo que ya esta dentro del alcance, cuanto lo ejecuta el equipo.")
    _tabla(ws, CE["eficacia"], mes_obj, pct_cols=("% Centralizacion",),
           formulas={
               "Total en Alcance": lambda r: (
                   f"={r('Ops Centrales')}+{r('Ops en Tienda (ya incorporadas)')}"),
               "% Centralizacion": lambda r: (
                   f"=IF({r('Total en Alcance')}=0,\"\","
                   f"{r('Ops Centrales')}/{r('Total en Alcance')})"),
           })
    ws.append([])

    _titulo(ws, "K6b · COBERTURA DE DESPLIEGUE", 6)
    _nota(ws, "Cuantas agencias del universo ya se incorporaron. ESTO es lo que de verdad crece con el "
              "tiempo. Para Pago Lote, el universo son las agencias CON CONVENIO, no toda la red.")
    _tabla(ws, CE["cobertura"], mes_obj, pct_cols=("% Cobertura",),
           formulas={
               "% Cobertura": lambda r: (
                   f"=IF({r('Universo')}=0,\"\",{r('Tiendas Incorporadas')}/{r('Universo')})"),
           })
    ws.append([])

    _titulo(ws, "K6c · FUGA RESIDUAL ACTIVA (agencias ya incorporadas)", 8)
    _nota(ws, "Operaciones que la agencia sigue haciendo POR SU CUENTA pese a estar ya incorporada al "
              "proceso. Cuando la cobertura (K6b) llega al 100%, este es el unico de los tres que queda "
              "vivo como KPI: es donde esta la accion.")
    _nota(ws, "'Tendencia' compara el ultimo mes contra el anterior: SUBE = la agencia esta retomando "
              "trabajo por su cuenta (revisar por que); 'nueva' = la familia recien entro en produccion.")
    if CE.get("fuga"):
        _tabla(ws, CE["fuga"], mes_obj, money_cols=("Monto (S/)",))
    else:
        _nota(ws, "Sin fuga activa en el ultimo mes.")
    ws.append([])

    if CE.get("fuga_cerradas"):
        _titulo(ws, "FUGA YA CERRADA (historico, sin actividad en el ultimo mes)", 4)
        _nota(ws, "Agencias que dejaron de operar por su cuenta: evidencia de que la centralizacion "
                  "efectivamente se consolido ahi.")
        _tabla(ws, CE["fuga_cerradas"], mes_obj, money_cols=("Monto Historico (S/)",))
        ws.append([])

    if CE.get("anomalias"):
        _titulo(ws, "ANOMALIAS — opera sin pertenecer al universo", 5)
        _nota(ws, "Agencias que registran una familia sin pertenecer a su universo (ej. procesar Pago "
                  "Lote sin tener convenio). Se apartan del denominador para no ensuciar los "
                  "porcentajes, pero se reportan aca: casi siempre indican un error de registro.")
        _tabla(ws, CE["anomalias"], mes_obj, money_cols=("Monto (S/)",))


def _sec_buzon(ws, K, mes_obj):
    _corte(ws)
    _titulo(ws, "BUZON DE REQUERIMIENTOS", 3)
    _nota(ws, "Categorias tal como se escriben en el CSV: variantes de tipeo pueden aparecer separadas.")
    if K["buzon_tipo"]:
        _tabla(ws, K["buzon_tipo"], None, pct_cols=("% del Total",))
        ws.append([])
    if K["buzon_asig"]:
        _tabla(ws, K["buzon_asig"], None, pct_cols=("% del Total",))
        ws.append([])
    if K["buzon_alerta"]:
        ws.append([K["buzon_alerta"]])
        c = ws.cell(ws.max_row, 1)
        c.font = Font(bold=True, color="9C0006")
        c.fill = PatternFill("solid", fgColor="FFC7CE")


def _hoja_explicacion(wb, K, mes_obj):
    ws = wb.create_sheet("Explicacion")
    _titulo(ws, "METODOLOGIA Y EXPLICACION DE CADA KPI", 4)
    explicacion = [
        ("KPI", "Que mide", "Como se calcula", "Fuente"),
        ("K1 Volumen Pago Lote", "Cantidad de transacciones de pago masivo (planillas)",
         "Conteo de filas 'PAGO LOTE' por mes", "Detalle Of. Principal"),
        ("K2 Monto Pago Lote", "Soles movilizados en pago lote",
         "Suma de 'Monto' de las filas Pago Lote", "Detalle Of. Principal"),
        ("K3 Volumen CTS", "Cantidad de depositos lote de CTS",
         "Conteo de filas 'DEPOSITO LOTE CTS' por mes", "Detalle Of. Principal"),
        ("K4 Monto CTS", "Soles depositados por CTS",
         "Suma de 'Monto' de las filas CTS", "Detalle Of. Principal"),
        ("K5 Ratio carga Asistentes", "Desbalance de carga entre asistentes",
         "MAX(ops asistente) / MIN(ops asistente). Meta <= 2.0x", "Detalle Of. Principal - Usuario"),
        ("K5a % Analista operando", "Cuanto del volumen lo hace el analista (deberia delegar)",
         "Ops del analista / Total ops. Meta <= 10%", "Detalle Of. Principal - Usuario"),
        ("K6a Eficacia de Centralizacion", "De lo que ya esta dentro del alcance, cuanto lo ejecuta el equipo",
         "Ops de Oficina Principal / (Ops Of. Principal + Ops de agencias YA incorporadas), por familia y "
         "desde que el proceso entro en produccion.", "Detalle Of. Principal + Detalle Tiendas"),
        ("K6b Cobertura de Despliegue", "Cuantas agencias del universo ya se incorporaron al proceso",
         "Agencias incorporadas (acumulado) / universo de la familia. Para Pago Lote el universo son "
         "las agencias CON CONVENIO, no toda la red.", "Detalle Of. Principal + Detalle Tiendas"),
        ("K6c Fuga Residual", "Operaciones que la agencia sigue haciendo por su cuenta pese a estar incorporada",
         "Ops en Tiendas con fecha posterior al alta de esa agencia. Se listan solo las que siguen "
         "activas en el ultimo mes; las que ya se cerraron van resumidas aparte.",
         "Detalle Tiendas + fecha de alta por agencia"),
        ("Anomalias de universo", "Agencias que operan una familia sin pertenecer a su universo",
         "Ops en Tiendas de agencias listadas en FUERA_UNIVERSO_POR_FAMILIA (ej. Pago Lote sin "
         "convenio). Se apartan del denominador y se reportan aparte: suelen ser error de registro.",
         "Detalle Tiendas"),
        ("K7 Eficiencia (hoja)", "Productividad por operador",
         "Total ops / persona-dias (incluye personal de apoyo) y / (operadores x dias habiles reales del mes)",
         "Detalle Of. Principal"),
        ("Dia Pico / Dias Pico en el Mes", "Carga laboral: que dia(s) concentran mas operaciones",
         "Dia con mas ops del mes; dias que superan promedio + 1 desv. estandar del mes.", "Detalle Of. Principal - Fecha"),
        ("K8 Tasa Extornos (hoja)", "Calidad: errores que se revierten",
         "Extornos / TOTAL de operaciones del mes (Pago Lote + CTS + aperturas + pendientes).", "Extornos"),
        ("K9 Aperturas y Abonos (hoja)", "Volumen de aperturas de CTS/Sueldo y de abonos de sueldo del equipo",
         "Aperturas: conteo por familia del Detalle. Abonos: conteo del reporte de Abono filtrado al equipo.",
         "Detalle Of. Principal + Abono de Sueldos y CTS"),
        ("Empleadores CTS", "Cantidad de empresas atendidas en CTS",
         "Conteo de empleadores unicos del mes", "Detalle CTS Individual - Empleador"),
        ("Tiempo Proc. (auxiliar)", "Carga horaria estimada del area (no es un KPI numerado)",
         f"K1x{MIN_POR_PAGO_LOTE}min + K3x{MIN_POR_CTS}min (ESTIMADO)", "Detalle Of. Principal"),
        ("Buzon", "Demanda de requerimientos y su reparto",
         "Conteo por tipo de solicitud y por persona asignada", "REQUERIMIENTOS ...csv"),
    ]
    for i, fila in enumerate(explicacion):
        ws.append(list(fila))
        r = ws.max_row
        for j in range(4):
            c = ws.cell(r, j + 1)
            c.border = BORDE
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 0:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=AZUL)
            elif r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GRIS)
    for j, w in enumerate((26, 42, 50, 34)):
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.append([])
    _nota(ws, "SUPUESTOS: operadores y roles definidos en ROLES; tiempos de procesamiento son estimados; "
              "el universo de Pago Lote son las agencias con convenio (se auto-detecta de la data).")
    _nota(ws, "LIMITACIONES CONOCIDAS: no hay motivo de extorno estandarizado; el buzon no siempre "
              "registra fecha de atencion; el tiempo de procesamiento no esta cronometrado.")


def escribir_excel(K, salida, mes_obj):
    """
    Arma el libro: una hoja de CALCULOS con todas las tablas apiladas, una de
    GRAFICOS y una de EXPLICACION. Cada seccion vive en su propia funcion.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculos"

    _sec_resumen(ws, K, mes_obj)
    _sec_operadores(ws, K, mes_obj)
    _sec_eficiencia(ws, K, mes_obj)
    _sec_extornos(ws, K, mes_obj)
    _sec_k9(ws, K, mes_obj)
    _sec_centralizacion(ws, K, mes_obj)
    _sec_buzon(ws, K, mes_obj)

    _hoja_graficos(wb, K)
    _hoja_explicacion(wb, K, mes_obj)

    for hoja in wb.worksheets:
        hoja.sheet_view.showGridLines = False
    wb.save(salida)


# ---------------------------------------------------------------------------
# MAIN -----------------------------------------------------------------------
# ---------------------------------------------------------------------------
def main():
    print("=" * 64)
    print("  GENERADOR DE KPIs — PAGOS MASIVOS Y CTS")
    print("=" * 64)

    f_det, f_tie = buscar_detalle()
    f_cts = buscar_archivo("cts", "individual")
    f_ext = buscar_archivos("extorno")
    f_abo = buscar_archivo("abono", "sueldo")
    f_req = buscar_archivo("requerimiento", ext=("csv",))

    for etq, ruta in [("Detalle Of. Principal", f_det), ("Detalle Tiendas", f_tie),
                      ("Detalle CTS", f_cts), ("Abono Sueldos", f_abo),
                      ("Requerimientos", f_req)]:
        print(f"  [{'OK ' if ruta else '-- '}] {etq:20s}: {os.path.basename(ruta) if ruta else 'NO ENCONTRADO'}")
    print(f"  [{'OK ' if f_ext else '-- '}] {'Extornos':20s}: "
          + (f"{len(f_ext)} archivo(s)" if f_ext else "NO ENCONTRADO"))

    if not f_det:
        print("\n  ERROR: no encuentro 'Detalle de Pago Lote y CTS' de OFICINA")
        print("  PRINCIPAL. Es obligatorio: es la fuente de todos los KPIs del equipo.")
        print("  (Corre 'python generar_datos_demo.py' si aun no tienes reportes de prueba.)")
        return
    if not f_tie:
        print("\n  AVISO: falta 'Detalle de Pago Lote y CTS - Tiendas'. Sin el no se")
        print("  puede calcular centralizacion ni fuga residual (K6).")

    print("\n  Leyendo y calculando...")
    det = leer_detalle_centrales(f_det, origen="CENTRALES")
    det_tie = leer_detalle_centrales(f_tie, origen="TIENDAS") if f_tie else pd.DataFrame()
    cts = leer_detalle_cts(f_cts) if f_cts else pd.DataFrame()
    ext = leer_extornos(f_ext) if f_ext else pd.DataFrame()
    abo = leer_abono(f_abo) if f_abo else pd.DataFrame()
    req = leer_requerimientos(f_req) if f_req else pd.DataFrame()

    meses_detectados = sorted(det["mes"].dropna().unique().tolist()) if not det.empty else []
    if not meses_detectados:
        print("  ERROR: no se detectaron meses con datos."); return

    corte = fecha_proceso(f_det)
    meses_info = clasificar_meses(det, corte)
    print(f"\n  Corte del reporte: {corte.strftime('%d/%m/%Y') if corte else 'no detectado'}")
    for mes, i in meses_info.items():
        marca = "OK" if i["comparable"] else "!!"
        print(f"  [{marca}] {mes}  {i['clase']:9s} "
              f"{i['dias_actividad']:2d}/{i['dias_habiles']:2d} dias con actividad"
              + (f"  -> {i['nota']}" if i["nota"] else ""))
    no_comp = [m for m, i in meses_info.items() if not i["comparable"]]
    if no_comp:
        print(f"  -> excluidos del pico de referencia de K7: {', '.join(no_comp)}")

    historico = cargar_historico()
    meses_cache = historico.get("meses", []) if historico else []
    print()
    if meses_cache:
        print(f"  Historico guardado: {', '.join(meses_cache)}")
        usar_historico = preguntar_si_no(
            "Conservar el historico ya calculado y actualizar SOLO los meses nuevos")
    else:
        usar_historico = preguntar_si_no(
            "Activar modo historico (estos meses quedaran fijos; en proximas corridas "
            "solo se agregaran meses nuevos)")

    if usar_historico and meses_cache:
        meses_nuevos = [m for m in meses_detectados if m not in meses_cache]
        if not meses_nuevos:
            print("  No hay meses nuevos: se usa el historico tal cual, sin recalcular.")
            K = reconstituir_desde_historico(historico)
        else:
            print(f"  Meses nuevos a calcular: {', '.join(meses_nuevos)} (el resto queda intacto)")
            K_nuevo = calcular(det, cts, ext, abo, meses_incluir=meses_nuevos, meses_info=meses_info)
            K = fusionar_historico(historico, K_nuevo)
            guardar_historico(K, meses_info)
            print(f"  Historico actualizado: {os.path.basename(CACHE_HISTORICO)}")
    else:
        K = calcular(det, cts, ext, abo, meses_info=meses_info)
        K["eficiencia"] = sintetizar_eficiencia(K["eficiencia"])
        if usar_historico:
            guardar_historico(K, meses_info)
            print(f"  Se creo un historico nuevo: {os.path.basename(CACHE_HISTORICO)}")

    if not K["meses"]:
        print("  ERROR: no se detectaron meses con datos."); return

    K["buzon_tipo"], K["buzon_asig"], K["buzon_alerta"] = calcular_buzon(req)
    K["carga_diaria"] = calcular_carga_diaria(det)
    K["centralizacion_real"] = calcular_centralizacion(det, det_tie)
    fusionar_absorcion_en_resumen(K)

    mes_obj = MES_OBJETIVO or K["meses"][-1]
    salida = os.path.join(CARPETA, f"KPIs_Pagos_CTS_{mes_obj}.xlsx")
    escribir_excel(K, salida, mes_obj)

    print(f"  Meses detectados: {', '.join(K['meses'])}")
    print(f"  Mes destacado   : {mes_obj}")
    fila = next((r for r in K["resumen"] if r["Mes"] == mes_obj), None)
    if fila:
        print("\n  Resumen del mes destacado:")
        print(f"    K1 Vol Pago Lote : {fila['K1 Vol Pago']:>8}   (S/ {fila['K2 Monto Pago']:,.2f})")
        print(f"    K3 Vol CTS       : {fila['K3 Vol CTS']:>8}   (S/ {fila['K4 Monto CTS']:,.2f})")
        print(f"    Total Ops        : {fila['Total Ops']:>8}   (S/ {fila['Total Monto']:,.2f})")
        print(f"    Empleadores CTS  : {fila['Empleadores CTS']:>8}")
    ek = next((e for e in K["extornos"] if e["Mes"] == mes_obj), None)
    if ek:
        print(f"    Extornos totales : {ek['Extornos Totales']:>8}   (imputables: {ek['Imputables (Error Usuario)']})")
    CE = K.get("centralizacion_real") or {}
    pl = [c for c in CE.get("eficacia", []) if c["Familia"] == "Pago Lote" and c["Mes"] == mes_obj]
    cob = [c for c in CE.get("cobertura", []) if c["Familia"] == "Pago Lote" and c["Mes"] == mes_obj]
    if pl:
        print(f"\n  K6 Centralizacion Pago Lote ({mes_obj}):")
        print(f"    Eficacia (en alcance): {pl[0]['% Centralizacion']*100:>6.1f}%   "
              f"({pl[0]['Ops Centrales']} equipo vs {pl[0]['Ops en Tienda (ya incorporadas)']} en agencia)")
    if cob:
        print(f"    Cobertura despliegue : {cob[0]['% Cobertura']*100:>6.1f}%   "
              f"({cob[0]['Tiendas Incorporadas']}/{cob[0]['Universo']} agencias con convenio)")
    print(f"\n  Excel generado: {os.path.basename(salida)}")
    print("=" * 64)


if __name__ == "__main__":
    main()
